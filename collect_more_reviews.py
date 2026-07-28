"""
collect_more_reviews.py
-----------------------
Phase 2b — Extended Review Collection (v2)

Improvements over v1:
  • Scrolls page + clicks "Load More" buttons up to SCROLL_ROUNDS times
    before scraping — this forces HexaHealth (and similar SPAs) to render
    reviews that are lazy-loaded and invisible on first page load.
  • Tries SECONDARY sources (Practo, JustDial) for every hospital — even
    those whose primary URL is HexaHealth.  New reviews from a different
    site are added alongside the existing ones.
  • Also retries hospitals with "No Reviews" or "Scraper Error" status —
    with proper scrolling they often yield results this time.
  • Deduplicates by (hospital, reviewer_name) across all sources so no
    duplicate rows are ever written.

Run:
    python collect_more_reviews.py

Resume after crash:
    Just run again — checkpoint resumes automatically.
"""

import re
import time
import random
import requests
from pathlib import Path
from urllib.parse import quote_plus
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from config import HEADLESS, SEARCH_DELAY_MIN, SEARCH_DELAY_MAX
from logger import get_logger
from checkpoint import save_checkpoint, load_checkpoint, clear_checkpoint
from excel_writer import ExcelWriter
from review_scraper import scrape_reviews          # HexaHealth scraper

log = get_logger("MORE_REVIEWS_V2")

EXCEL_FILE = Path("output/HospitalAutomation.xlsx")

# How many reviews to scrape per hospital per source.
MAX_SCRAPE_PER_HOSPITAL = 50

# How many times to scroll / click "Load More" before reading the HTML.
# Each round scrolls to the bottom and waits 2 s.  On HexaHealth this
# typically loads an extra 5–10 reviews per round.
SCROLL_ROUNDS = 6

# Delay between hospitals (seconds)
MIN_DELAY = 1.5
MAX_DELAY = 3.0

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/138.0 Safari/537.36"
    )
}

# Statuses eligible for re-scraping
ELIGIBLE_STATUSES = {"Success", "Low Reviews", "No Reviews", "Scraper Error"}

# Secondary sites to try for every hospital regardless of primary source
SECONDARY_SITES = ["hexahealth.com", "practo.com", "justdial.com"]


# ─────────────────────────────────────────────────────────────────────────────
# Scroll helper
# ─────────────────────────────────────────────────────────────────────────────

def scroll_and_load(page, rounds: int = SCROLL_ROUNDS):
    """
    Scroll to the bottom of the page ROUNDS times, clicking any visible
    'Load More' / 'Show More' / 'See All' buttons along the way.
    Waits 2 s after each scroll to let the SPA fetch and render new content.
    """
    load_more_selectors = [
        "button:has-text('Load More')",
        "button:has-text('Show More')",
        "button:has-text('See All Reviews')",
        "button:has-text('View More')",
        "a:has-text('Load More')",
        "a:has-text('Show More')",
        "span:has-text('Load More')",
        "[class*='loadMore']",
        "[class*='load-more']",
        "[class*='showMore']",
        "[class*='show-more']",
        "[class*='viewMore']",
    ]

    for round_num in range(1, rounds + 1):
        # Scroll to bottom
        page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
        page.wait_for_timeout(1_800)

        # Try clicking any "load more"-style button
        clicked = False
        for selector in load_more_selectors:
            try:
                btn = page.locator(selector).first
                if btn.count() and btn.is_visible():
                    btn.scroll_into_view_if_needed()
                    btn.click()
                    page.wait_for_timeout(1_800)
                    clicked = True
                    log.info(f"  Clicked '{selector}' on round {round_num}")
                    break
            except Exception:
                pass

        if not clicked:
            # Nothing to click — extra scroll just in case
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(1_000)


# ─────────────────────────────────────────────────────────────────────────────
# Excel loaders
# ─────────────────────────────────────────────────────────────────────────────

def load_hospitals_to_extend() -> list[dict]:
    """
    Read the Summary sheet and return all hospitals whose status is in
    ELIGIBLE_STATUSES.  Hospitals with a known URL are used directly;
    those without (No Reviews / Scraper Error) will be re-searched.
    """
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"{EXCEL_FILE} not found. Run collect_reviews.py first."
        )

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Summary"]

    hospitals = []
    seen = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        name   = row[0]
        status = row[1]
        source = row[4]
        url    = row[5]

        if not name:
            continue
        status_str = str(status or "").strip()
        if status_str not in ELIGIBLE_STATUSES:
            continue
        name_str = str(name).strip()
        if name_str in seen:
            continue

        seen.add(name_str)
        hospitals.append({
            "name":   name_str,
            "status": status_str,
            "source": str(source or "hexahealth.com").strip(),
            "url":    str(url or "").strip(),
        })

    wb.close()
    log.info(f"Found {len(hospitals)} hospital(s) eligible for extended scraping.")
    return hospitals


def load_existing_reviews() -> dict:
    """
    Return {hospital_name: set(reviewer_names_lower)} from the Reviews sheet.
    """
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Reviews"]

    existing: dict = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        hospital = row[0]
        reviewer = row[1]
        if hospital and reviewer:
            key = str(hospital).strip()
            existing.setdefault(key, set()).add(str(reviewer).strip().lower())

    wb.close()
    return existing


# ─────────────────────────────────────────────────────────────────────────────
# URL helpers (same logic as review_finder.py)
# ─────────────────────────────────────────────────────────────────────────────

def _fix_hexahealth_url(url: str) -> str:
    match = re.match(
        r"(https?://[^/]+/[^/]+/(?:hospital|doctor)/[^/?#]+)", url
    )
    if match:
        return match.group(1).rstrip("/") + "/reviews"
    if "/reviews" not in url.lower():
        return url.rstrip("/") + "/reviews"
    return url


def _duckduckgo_search(page, hospital_name: str, site: str) -> str | None:
    query = quote_plus(f"{hospital_name} reviews site:{site}")
    ddg_url = f"https://duckduckgo.com/?q={query}&ia=web"
    try:
        page.goto(ddg_url, wait_until="domcontentloaded", timeout=30_000)
        page.wait_for_timeout(2_500)
        for anchor in page.locator("a[href]").all():
            href = (anchor.get_attribute("href") or "").strip()
            if (
                href.startswith("http")
                and site in href
                and "duckduckgo.com" not in href
            ):
                return href
    except Exception as exc:
        log.warning(f"  DDG error ({site}): {exc}")
    return None


def find_url_for_site(page, hospital_name: str, site: str) -> str | None:
    """Search DuckDuckGo for the hospital on a specific site and return URL."""
    url = _duckduckgo_search(page, hospital_name, site)
    if not url:
        return None
    if site == "hexahealth.com":
        url = _fix_hexahealth_url(url)
    return url


# ─────────────────────────────────────────────────────────────────────────────
# Site-specific scrapers
# ─────────────────────────────────────────────────────────────────────────────

def _scrape_practo(url: str, max_reviews: int, page=None) -> list[dict]:
    if page is not None:
        html = page.content()
    else:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        html = resp.text
    soup = BeautifulSoup(html, "html.parser")

    cards = (
        soup.select("div.doctor-review-text")
        or soup.select("div[class*='review-card']")
        or soup.select("div[class*='review_card']")
        or soup.select("div[class*='ReviewCard']")
        or soup.select("div[data-qa-id='review_card']")
    )

    reviews = []
    for card in cards:
        name_tag = (
            card.find("p",    class_=re.compile(r"reviewer|author|username", re.I))
            or card.find("span", class_=re.compile(r"reviewer|author|username", re.I))
            or card.find("strong")
        )
        reviewer = name_tag.get_text(strip=True) if name_tag else "Unknown"

        stars = card.select("i.fa-star, span.fa-star, img[alt*='star']")
        rating = len(stars) if stars else 0
        if rating == 0:
            rating_tag = card.find(class_=re.compile(r"rating|score", re.I))
            if rating_tag:
                try:
                    rating = float(re.search(r"\d+\.?\d*", rating_tag.get_text()).group())
                except Exception:
                    rating = 0

        text_tag = (
            card.find("p",   class_=re.compile(r"review.?text|comment|body", re.I))
            or card.find("div", class_=re.compile(r"review.?text|comment|body", re.I))
            or card.find("p")
        )
        review_text = text_tag.get_text(strip=True) if text_tag else ""

        date_tag = card.find(class_=re.compile(r"date|time|when", re.I))
        date = date_tag.get_text(strip=True) if date_tag else ""

        reviews.append({"reviewer": reviewer, "rating": rating,
                        "review": review_text, "date": date})
        if len(reviews) >= max_reviews:
            break
    return reviews


def _scrape_justdial(url: str, max_reviews: int, page=None) -> list[dict]:
    if page is not None:
        html = page.content()
    else:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        html = resp.text
    soup = BeautifulSoup(html, "html.parser")

    cards = (
        soup.select("div.review-box")
        or soup.select("div[class*='reviewdesc']")
        or soup.select("div[class*='review_box']")
        or soup.select("div[class*='ReviewBox']")
        or soup.select("li.review-item")
    )

    reviews = []
    for card in cards:
        name_tag = (
            card.find(class_=re.compile(r"reviewer|user.?name|author", re.I))
            or card.find("span")
        )
        reviewer = name_tag.get_text(strip=True) if name_tag else "Unknown"

        stars = card.select("i.icon-star-full, span[class*='star'], em[class*='star']")
        rating = len(stars) if stars else 0
        if rating == 0:
            rating_tag = card.find(class_=re.compile(r"rating|score|star", re.I))
            if rating_tag:
                try:
                    rating = float(re.search(r"\d+\.?\d*", rating_tag.get_text()).group())
                except Exception:
                    rating = 0

        text_tag = (
            card.find("p",   class_=re.compile(r"review.?text|comment|desc", re.I))
            or card.find("div", class_=re.compile(r"review.?text|comment|desc", re.I))
            or card.find("p")
        )
        review_text = text_tag.get_text(strip=True) if text_tag else ""

        date_tag = card.find(class_=re.compile(r"date|time|posted", re.I))
        date = date_tag.get_text(strip=True) if date_tag else ""

        reviews.append({"reviewer": reviewer, "rating": rating,
                        "review": review_text, "date": date})
        if len(reviews) >= max_reviews:
            break
    return reviews


def scrape_for_site(site: str, url: str, max_reviews: int, page=None) -> list[dict]:
    if site == "hexahealth.com":
        return scrape_reviews(url, max_reviews=max_reviews, page=page)
    elif site == "practo.com":
        return _scrape_practo(url, max_reviews=max_reviews, page=page)
    elif site == "justdial.com":
        return _scrape_justdial(url, max_reviews=max_reviews, page=page)
    else:
        return scrape_reviews(url, max_reviews=max_reviews, page=page)


# ─────────────────────────────────────────────────────────────────────────────
# Core: scrape one (hospital, site, url) combination
# ─────────────────────────────────────────────────────────────────────────────

def scrape_one_source(
    page,
    hospital_name: str,
    site: str,
    url: str,
    already_have: set,
) -> list[dict]:
    """
    Navigate → scroll/load → scrape → deduplicate.
    Returns list of NEW review dicts.  Updates already_have in-place.
    """
    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30_000)
        page.wait_for_timeout(2_000)
    except Exception as exc:
        log.error(f"  Could not load {url}: {exc}")
        return []

    # ── Scroll & click Load More ─────────────────────────────────────────────
    scroll_and_load(page, rounds=SCROLL_ROUNDS)

    # ── Scrape ───────────────────────────────────────────────────────────────
    try:
        all_reviews = scrape_for_site(site, url, max_reviews=MAX_SCRAPE_PER_HOSPITAL, page=page)
    except Exception as exc:
        log.error(f"  Scraper error on {site} for {hospital_name}: {exc}")
        return []

    # ── Deduplicate ──────────────────────────────────────────────────────────
    new_reviews = []
    seen_this_batch: set = set()
    for r in all_reviews:
        key = str(r.get("reviewer", "Unknown")).strip().lower()
        if key in already_have or key in seen_this_batch:
            continue
        new_reviews.append(r)
        already_have.add(key)
        seen_this_batch.add(key)

    print(f"    [{site}] scraped={len(all_reviews)}  new={len(new_reviews)}")
    return new_reviews


# ─────────────────────────────────────────────────────────────────────────────
# Process one hospital across all sources
# ─────────────────────────────────────────────────────────────────────────────

def process_hospital(
    page,
    hospital: dict,
    existing_reviews: dict,
    writer: ExcelWriter,
) -> int:
    name         = hospital["name"]
    primary_site = hospital["source"]
    primary_url  = hospital["url"]

    already_have = existing_reviews.get(name, set()).copy()
    # update existing_reviews in place so checkpoint saves carry it forward
    existing_reviews[name] = already_have

    print(f"\n{'='*70}")
    print(f"Hospital : {name}  [{hospital['status']}]")
    print(f"Primary  : {primary_site}  |  Existing reviews: {len(already_have)}")

    total_new = 0

    # ── 1. Primary source (known URL — no search needed) ────────────────────
    if primary_url:
        print(f"  Primary URL: {primary_url}")
        new = scrape_one_source(page, name, primary_site, primary_url, already_have)
        for r in new:
            writer.add_review(
                hospital_name=name,
                reviewer=r.get("reviewer", "Unknown"),
                rating=r.get("rating", ""),
                review=r.get("review", ""),
                review_date=r.get("date", ""),
                source=primary_site,
                upload_status="Pending",
            )
        total_new += len(new)
        time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

    # ── 2. Secondary sources (DuckDuckGo search → scrape) ───────────────────
    for site in SECONDARY_SITES:
        if site == primary_site:
            continue   # already done above

        print(f"  Searching secondary: {site} …")
        try:
            sec_url = find_url_for_site(page, name, site)
        except Exception as exc:
            log.warning(f"  Search error ({site}): {exc}")
            sec_url = None

        if not sec_url:
            print(f"    [{site}] no URL found — skipping")
            time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))
            continue

        print(f"    [{site}] URL: {sec_url}")
        new = scrape_one_source(page, name, site, sec_url, already_have)
        for r in new:
            writer.add_review(
                hospital_name=name,
                reviewer=r.get("reviewer", "Unknown"),
                rating=r.get("rating", ""),
                review=r.get("review", ""),
                review_date=r.get("date", ""),
                source=site,
                upload_status="Pending",
            )
        total_new += len(new)
        time.sleep(random.uniform(MIN_DELAY, MAX_DELAY))

    print(f"  → {total_new} new review(s) added for {name}")
    log.info(f"process_hospital done: {name} → +{total_new}")
    return total_new


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    hospitals = load_hospitals_to_extend()
    total     = len(hospitals)

    if total == 0:
        print("No eligible hospitals found in Summary sheet.")
        return

    # ── Resume support ───────────────────────────────────────────────────────
    checkpoint = load_checkpoint()
    if checkpoint and checkpoint.get("phase") == "collect_more_v2":
        start_index = checkpoint["hospital_index"]
        print(f"\nResuming from hospital #{start_index}: {checkpoint['hospital_name']}")
        log.info(f"Resuming collect_more_v2 from index {start_index}")
    else:
        start_index = 0
        print(f"\nStarting extended collection (v2) for {total} hospital(s).")
        print(f"Strategy: primary URL + up to {len(SECONDARY_SITES)-1} secondary sites, "
              f"{SCROLL_ROUNDS} scroll rounds each, up to {MAX_SCRAPE_PER_HOSPITAL} reviews/source.")

    existing_reviews = load_existing_reviews()
    log.info(f"Loaded existing reviewers for {len(existing_reviews)} hospital(s).")

    writer    = ExcelWriter()
    total_new = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        page    = browser.new_page()

        try:
            for i, hospital in enumerate(
                hospitals[start_index:], start=start_index
            ):
                try:
                    added = process_hospital(page, hospital, existing_reviews, writer)
                    total_new += added
                except Exception as exc:
                    log.error(f"Unhandled error for {hospital['name']}: {exc}")
                    print(f"  ERROR: {exc} — skipping.")

                save_checkpoint(
                    phase="collect_more_v2",
                    hospital_index=i + 1,
                    hospital_name=hospital["name"],
                )
                print(f"  Progress: {i+1}/{total}  |  Total new so far: {total_new}")

        finally:
            browser.close()

    clear_checkpoint()
    print("\n" + "=" * 70)
    print("Extended collection v2 complete.")
    print(f"Total new reviews added : {total_new}")
    print(f"Results saved to        : {EXCEL_FILE}")
    log.info(f"collect_more_v2 finished — {total_new} new reviews across {total} hospitals.")


if __name__ == "__main__":
    main()
