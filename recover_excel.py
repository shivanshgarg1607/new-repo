"""
recover_excel.py
-----------------
Run this when the Excel file is corrupted ("Bad magic number for central
directory"). It tries multiple recovery strategies in order:

  1. Restore from automatic backup (.bak.xlsx)
  2. Restore from temp file (.tmp.xlsx)
  3. Repair the ZIP structure and extract via openpyxl
  4. Parse the raw XML inside the file to salvage cell data
  5. Rebuild from the root copy (HospitalAutomation.xlsx in project root)

Usage:  python recover_excel.py
"""

from pathlib import Path
import shutil
import sys
import re
import zipfile
import io

EXCEL_FILE = Path("output/HospitalAutomation.xlsx")
BAK_FILE   = Path("output/HospitalAutomation.bak.xlsx")
TMP_FILE   = Path("output/HospitalAutomation.tmp.xlsx")
ROOT_FILE  = Path("HospitalAutomation.xlsx")

from openpyxl import load_workbook, Workbook


def try_load(path):
    try:
        wb = load_workbook(path)
        wb.close()
        return True
    except Exception:
        return False


# ─────────────────────────────────────────────
# Strategy 3: repair the ZIP central directory
# ─────────────────────────────────────────────

def try_repair_zip(path):
    """
    An .xlsx is a ZIP archive. "Bad magic number for central directory"
    means the central directory at the end of the file is damaged, but
    the individual files inside may still be intact. We rebuild the ZIP
    by scanning for local file headers and re-archiving them.
    """
    raw = path.read_bytes()
    if not raw:
        return None

    # Find all local file headers: PK\x03\x04
    local_sig = b"PK\x03\x04"
    central_sig = b"PK\x01\x02"

    offsets = []
    idx = 0
    while True:
        pos = raw.find(local_sig, idx)
        if pos == -1:
            break
        offsets.append(pos)
        idx = pos + 4

    if not offsets:
        return None

    # Try to extract each entry by finding the next local header or
    # central directory header as the boundary.
    entries = []
    for i, start in enumerate(offsets):
        end = offsets[i + 1] if i + 1 < len(offsets) else len(raw)
        # But there might be a central directory after the last local entry
        central_pos = raw.find(central_sig, start + 4, end)
        if central_pos != -1 and central_pos < end:
            end = central_pos
        chunk = raw[start:end]
        entries.append(chunk)

    if not entries:
        return None

    # Rebuild a proper ZIP from the extracted chunks
    new_zip = io.BytesIO()
    with zipfile.ZipFile(new_zip, "w", zipfile.ZIP_DEFLATED) as zf:
        for chunk in entries:
            # Parse the local header to get filename and data
            if len(chunk) < 30:
                continue
            try:
                fname_len = int.from_bytes(chunk[26:28], "little")
                extra_len = int.from_bytes(chunk[28:30], "little")
                comp_method = int.from_bytes(chunk[8:10], "little")
                comp_size = int.from_bytes(chunk[18:22], "little")
                fname = chunk[30:30 + fname_len].decode("utf-8", errors="replace")
                data_start = 30 + fname_len + extra_len
                file_data = chunk[data_start:data_start + comp_size]

                if comp_method == 0:
                    raw_data = file_data
                elif comp_method == 8:
                    import zlib
                    try:
                        raw_data = zlib.decompress(file_data, -15)
                    except Exception:
                        continue
                else:
                    continue

                zf.writestr(fname, raw_data)
            except Exception:
                continue

    new_zip.seek(0)
    return new_zip


# ─────────────────────────────────────────────
# Strategy 4: parse raw XML to salvage cell data
# ─────────────────────────────────────────────

def salvage_from_xml(path):
    """
    Last resort: scan the raw bytes for <sheetData> XML blocks and
    extract cell values directly. This won't preserve formatting but
    will recover the actual text data.
    """
    raw = path.read_bytes()
    if not raw:
        return None

    # Find all sheet XML files by looking for <sheetData> blocks
    sheet_data_pattern = rb"<sheetData>(.*?)</sheetData>"
    sheets = re.findall(sheet_data_pattern, raw, re.DOTALL)

    if not sheets:
        return None

    # Also find shared strings (used for text values)
    shared_strings = []
    sst_pattern = rb"<t[^>]*>(.*?)</t>"
    sst_matches = re.findall(sst_pattern, raw, re.DOTALL)
    # Filter to only those in the sharedStrings.xml portion
    sst_start = raw.find(b"sharedStrings")
    if sst_start != -1:
        sst_end = raw.find(b"PK", sst_start + 20)
        if sst_end == -1:
            sst_end = len(raw)
        sst_section = raw[sst_start:sst_end]
        sst_matches = re.findall(sst_pattern, sst_section, re.DOTALL)

    for match in sst_matches:
        text = match.decode("utf-8", errors="replace")
        shared_strings.append(text)

    # Parse rows from sheet data
    all_sheets_data = []
    row_pattern = rb'<row r="(\d+)"[^>]*>(.*?)</row>'
    cell_pattern = rb'<c r="([A-Z]+)(\d+)"(?:\s+s="(\d+)")?(?:\s+t="(\w+)")?\s*(?:/>|>(?:<v>(.*?)</v>)?(?:<is><t[^>]*>(.*?)</t></is>)?</c>)?'

    for sheet_xml in sheets:
        rows = {}
        for row_match in re.finditer(row_pattern, sheet_xml, re.DOTALL):
            row_num = int(row_match.group(1))
            row_content = row_match.group(2)

            cells = {}
            for cell_match in re.finditer(cell_pattern, row_content, re.DOTALL):
                col_letters = cell_match.group(1).decode()
                cell_type = cell_match.group(4)
                value = cell_match.group(5)
                inline_text = cell_match.group(6)

                if inline_text:
                    val = inline_text.decode("utf-8", errors="replace")
                elif value:
                    val = value.decode("utf-8", errors="replace")
                    if cell_type and cell_type == b"s" and val.isdigit():
                        idx = int(val)
                        if idx < len(shared_strings):
                            val = shared_strings[idx]
                else:
                    val = ""

                cells[col_letters] = val

            if cells:
                rows[row_num] = cells

        if rows:
            all_sheets_data.append(rows)

    if not all_sheets_data:
        return None

    return all_sheets_data


def col_letter_to_num(letters):
    num = 0
    for ch in letters:
        num = num * 26 + (ord(ch) - ord("A") + 1)
    return num


def rebuild_from_salvage(sheets_data, excel_path):
    """Build a new workbook from salvaged XML cell data."""
    wb = Workbook()
    wb.remove(wb.active)

    sheet_names = ["Hospitals", "Reviews", "Upload Status", "Summary"]

    for i, rows_dict in enumerate(sheets_data):
        if i < len(sheet_names):
            name = sheet_names[i]
        else:
            name = f"Sheet{i + 1}"

        ws = wb.create_sheet(name)

        for row_num in sorted(rows_dict.keys()):
            cells = rows_dict[row_num]
            for col_letters, value in cells.items():
                col_num = col_letter_to_num(col_letters)
                ws.cell(row=row_num, column=col_num, value=value)

    from excel_lock import save_workbook
    save_workbook(wb, excel_path)
    return True


# ─────────────────────────────────────────────
# Strategy 5: rebuild from root copy
# ─────────────────────────────────────────────

def rebuild_from_root(excel_path):
    """Copy the root HospitalAutomation.xlsx as a fresh starting point."""
    if not ROOT_FILE.exists():
        return False
    try:
        wb = load_workbook(ROOT_FILE)
        # Ensure output directory exists
        excel_path.parent.mkdir(parents=True, exist_ok=True)
        from excel_lock import save_workbook
        save_workbook(wb, excel_path)
        return True
    except Exception:
        return False


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    print(f"Recovery for: {EXCEL_FILE}")
    print()

    if not EXCEL_FILE.exists():
        print("  Excel file does not exist — nothing to recover.")
        sys.exit(1)

    # Case 1: the main file loads fine — nothing to do.
    if try_load(EXCEL_FILE):
        print("  File loads OK — no recovery needed.")
        return

    print("  [1] Main file is corrupted.")

    # Case 2: try the backup.
    if BAK_FILE.exists() and try_load(BAK_FILE):
        print(f"  [2] Backup found at {BAK_FILE} — restoring.")
        shutil.copy2(str(BAK_FILE), str(EXCEL_FILE))
        print("  Restored from backup.")
        return

    print("  [2] No backup available.")

    # Case 3: try the temp file.
    if TMP_FILE.exists() and try_load(TMP_FILE):
        print(f"  [3] Temp file found at {TMP_FILE} — restoring.")
        shutil.copy2(str(TMP_FILE), str(EXCEL_FILE))
        print("  Restored from temp file.")
        return

    print("  [3] No temp file available.")

    # Case 4: try to repair the ZIP structure.
    print("  [4] Attempting ZIP repair…")
    repaired = try_repair_zip(EXCEL_FILE)
    if repaired is not None:
        repaired_path = EXCEL_FILE.with_suffix(".repaired.xlsx")
        with open(repaired_path, "wb") as f:
            f.write(repaired.getvalue())
        if try_load(repaired_path):
            shutil.copy2(str(repaired_path), str(EXCEL_FILE))
            repaired_path.unlink(missing_ok=True)
            print("  Repaired ZIP structure — file restored.")
            return
        else:
            print("  ZIP repair did not produce a valid file.")
            repaired_path.unlink(missing_ok=True)
    else:
        print("  Could not find any ZIP data to repair.")

    # Case 5: salvage raw XML data from the corrupted file.
    print("  [5] Attempting raw XML salvage…")
    sheets_data = salvage_from_xml(EXCEL_FILE)
    if sheets_data:
        try:
            rebuild_from_salvage(sheets_data, EXCEL_FILE)
            print(f"  Salvaged {len(sheets_data)} sheet(s) from raw XML data.")
            return
        except Exception as e:
            print(f"  XML salvage failed: {e}")
    else:
        print("  No XML data could be extracted.")

    # Case 6: rebuild from the root copy.
    print("  [6] Attempting rebuild from root copy…")
    if rebuild_from_root(EXCEL_FILE):
        print(f"  Rebuilt from {ROOT_FILE}.")
        print("  Note: This is a fresh file. You will need to re-run the")
        print("  collector to repopulate reviews. The checkpoint will let")
        print("  it resume from the last completed hospital.")
        return

    print()
    print("  All recovery strategies failed.")
    print("  Please restore from a manual backup or re-run the collector.")
    sys.exit(1)


if __name__ == "__main__":
    main()
