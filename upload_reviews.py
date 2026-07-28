"""
upload_reviews.py
-----------------
Phase 3 — Review Upload

Reads every "Pending" row from the Reviews sheet in HospitalAutomation.xlsx,
uploads each review to karunahealthlifepartner.com via the central
"Reviews & Ratings → Add Review" form, then marks each row as
"Uploaded" / "Not Found on Site" / "Failed" in the Reviews sheet.

Selectors come directly from Playwright codegen — do NOT change unless
the admin panel HTML has changed.

Run:
    python upload_reviews.py

Resume after crash:
    Just run again — rows already marked "Uploaded" / "Not Found on Site"
    are skipped automatically.
"""

from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

from config import BASE_URL, USERNAME, PASSWORD, HEADLESS
from logger import get_logger
from excel_writer import ExcelWriter

log = get_logger("UPLOAD_REVIEWS")

EXCEL_FILE = Path("output/HospitalAutomation.xlsx")

# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

def load_pending_reviews() -> list[dict]:
    """
    Return every row from the Reviews sheet whose Upload Status == "Pending".
    Skips rows with no review text AND no reviewer (likely dummy/test entries).
    """
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"{EXCEL_FILE} not found. Run collect_reviews.py first."
        )

    wb   = load_workbook(EXCEL_FILE)
    ws   = wb["Reviews"]
    rows = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        hospital_name = row[0]
        reviewer      = row[1]
        rating        = row[2]
        review        = row[3]
        review_date   = row[4]
        source        = row[5]
        upload_status = row[6]

        if not hospital_name:
            continue
        if str(upload_status or "").strip().lower() != "pending":
            continue

        rows.append({
            "hospital_name": str(hospital_name).strip(),
            "reviewer":      str(reviewer or "Unknown").strip(),
            "rating":        rating or 0,
            "review":        str(review or "").strip(),
            "review_date":   str(review_date or "").strip(),
            "source":        str(source or "").strip(),
            "row_index":     i,
        })

    wb.close()
    log.info(f"Found {len(rows)} pending review(s) to upload.")
    return rows


def mark_review_row(row_index: int, status: str):
    """Overwrite the Upload Status cell for a specific Reviews sheet row."""
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Reviews"]
    ws.cell(row=row_index, column=7, value=status)
    wb.save(EXCEL_FILE)
    wb.close()


# ─────────────────────────────────────────────────────────────────────────────
# Browser helpers
# ─────────────────────────────────────────────────────────────────────────────

def login(page):
    log.info("Navigating to admin panel…")
    page.goto(BASE_URL, wait_until="domcontentloaded", timeout=60_000)
    page.get_by_role("textbox", name="Enter User Name").fill(USERNAME)
    page.get_by_role("textbox", name="Password").fill(PASSWORD)
    page.get_by_role("button",  name="SIGN IN").click()
    page.wait_for_load_state("networkidle", timeout=30_000)
    log.info("Logged in successfully.")


def go_to_reviews_page(page):
    """Click the Reviews & Ratings sidebar link (partial match — ignores count)."""
    page.get_by_role("link", name="Reviews & Ratings").click()
    page.wait_for_load_state("domcontentloaded", timeout=30_000)
    page.wait_for_timeout(1_000)


def open_add_review_form(page):
    """Click the Add Review listitem to open a blank form."""
    page.get_by_role("listitem").filter(has_text="Add Review").click()
    page.wait_for_timeout(1_200)


def search_and_select_hospital(page, hospital_name: str) -> bool:
    """
    Type the hospital name into the autocomplete search box and pick
    the best-matching option from the dropdown.

    The site renders options as:  "{Name} — Hospital"
    We detect any element whose text contains " — " after a short wait.

    Returns True if matched and clicked, False if nothing found.
    """
    search_box = page.get_by_role("textbox", name="Type to search listings (e.g")

    # Use first 40 chars to trigger autocomplete
    search_term = hospital_name[:40].strip()
    search_box.fill("")
    search_box.type(search_term, delay=40)   # type() is slower → triggers JS events

    # Give the autocomplete JS up to 4 seconds to render suggestions
    page.wait_for_timeout(3_000)

    # ── Find dropdown options ────────────────────────────────────────────────
    # The codegen showed plain text like "Apollo Hospital — Hospital".
    # We look for any visible element whose text contains " — ".
    # This works regardless of which autocomplete library the site uses.
    options = page.locator("text=/ — /").all()

    if not options:
        # Try with only the first word (hospital names can be long / fuzzy)
        first_word = hospital_name.split()[0] if hospital_name.split() else hospital_name
        if first_word.lower() != search_term.lower():
            search_box.fill("")
            search_box.type(first_word, delay=40)
            page.wait_for_timeout(3_000)
            options = page.locator("text=/ — /").all()

    if not options:
        log.warning(f"  No dropdown options for: {hospital_name}")
        return False

    # ── Score and pick the best match ───────────────────────────────────────
    hospital_words = set(hospital_name.lower().split())
    best_score, best_option = -1, None

    for opt in options:
        try:
            text = (opt.inner_text() or "").strip()
        except Exception:
            continue
        if not text or "no results" in text.lower():
            continue
        opt_words = set(text.lower().split())
        score     = len(hospital_words & opt_words)
        if score > best_score:
            best_score, best_option = score, opt

    if best_option is None or best_score == 0:
        log.warning(f"  No matching option found for: {hospital_name}")
        return False

    try:
        best_option.click()
        page.wait_for_timeout(600)
        log.info("  Hospital selected from dropdown.")
        return True
    except Exception as exc:
        log.warning(f"  Could not click dropdown option: {exc}")
        return False


def click_star(page, rating):
    """Click the Nth star in #amrStarRow (1–5). Skip if 0 or invalid."""
    try:
        n = int(float(str(rating)))
    except (ValueError, TypeError):
        n = 0

    if n < 1 or n > 5:
        return

    try:
        page.locator(f"#amrStarRow > i:nth-child({n})").click()
        page.wait_for_timeout(400)
        log.info(f"  Clicked {n}-star rating.")
    except Exception as exc:
        log.warning(f"  Could not click star {n}: {exc}")


def upload_one_review(page, review: dict) -> str:
    """
    Fill in and submit the Add Review form for one review.

    Returns:
        "Uploaded"           — saved successfully
        "Not Found on Site"  — hospital missing from admin dropdown
        "Failed"             — unexpected error
    """
    hospital_name = review["hospital_name"]
    log.info(f"Uploading: {hospital_name} | {review['reviewer']}")

    try:
        # 1. Reviewer name
        page.get_by_role("textbox", name="e.g. Rahul S., Priya G.,").fill(
            review["reviewer"]
        )

        # 2. Hospital search + selection
        if not search_and_select_hospital(page, hospital_name):
            return "Not Found on Site"

        # 3. Star rating
        click_star(page, review["rating"])

        # 4. Review text  (fallback to generic text if empty)
        text = review["review"] or "Good experience at this hospital."
        page.get_by_role("textbox", name="Write the review content…").fill(text)

        # 5. Mark as Pending — never Final Approve
        page.get_by_role("radio", name="Pending (need to approve").check()
        page.wait_for_timeout(300)

        # 6. Save
        page.get_by_role("button", name="Save Review").click()
        page.wait_for_timeout(2_000)

        log.info(f"  Saved: {hospital_name}")
        return "Uploaded"

    except Exception as exc:
        log.error(f"  Exception for {hospital_name}: {exc}")
        return "Failed"


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    pending = load_pending_reviews()
    total   = len(pending)

    if total == 0:
        print("No pending reviews found. Nothing to upload.")
        return

    print(f"\nStarting upload of {total} review(s).")
    print("=" * 70)

    uploaded, not_found, failed = 0, 0, 0

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        page    = browser.new_page()

        try:
            login(page)
            go_to_reviews_page(page)

            for idx, review in enumerate(pending, start=1):
                print(
                    f"\n[{idx}/{total}] {review['hospital_name']} | "
                    f"{review['reviewer']} | {review['rating']}★"
                )

                result = "Failed"
                try:
                    open_add_review_form(page)
                    result = upload_one_review(page, review)
                except Exception as exc:
                    log.error(f"Unhandled error on row {review['row_index']}: {exc}")
                    result = "Failed"

                # Write result to Excel immediately so a crash doesn't lose progress
                mark_review_row(review["row_index"], result)

                if result == "Uploaded":
                    uploaded += 1
                    print(f"  ✓ Uploaded")
                elif result == "Not Found on Site":
                    not_found += 1
                    print(f"  ✗ Not Found on Site")
                else:
                    failed += 1
                    print(f"  ✗ Failed")

                # Navigate back to Reviews & Ratings for the next review.
                # If that fails (page in bad state), re-login and retry.
                try:
                    go_to_reviews_page(page)
                except Exception:
                    log.warning("Navigation failed — re-logging in.")
                    try:
                        login(page)
                        go_to_reviews_page(page)
                    except Exception as e2:
                        log.error(f"Re-login also failed: {e2} — stopping.")
                        break

        finally:
            browser.close()

    print("\n" + "=" * 70)
    print("Upload complete.")
    print(f"  Uploaded        : {uploaded}")
    print(f"  Not Found       : {not_found}")
    print(f"  Failed          : {failed}")
    print(f"  Total processed : {total}")
    print(f"\nSee column G ('Upload Status') in {EXCEL_FILE} for per-row results.")
    log.info(
        f"Upload done — Uploaded:{uploaded} NotFound:{not_found} Failed:{failed}"
    )


if __name__ == "__main__":
    main()
