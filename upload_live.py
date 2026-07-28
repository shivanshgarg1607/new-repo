"""
upload_live.py
--------------
Live Upload Watcher — runs ALONGSIDE scraper_main.py in a second terminal.

How it works:
  Every POLL_INTERVAL seconds it checks the Reviews sheet for rows where
  Upload Status is blank (newly scraped by scraper_main.py).
  It immediately claims those rows (marks them "Pending"), then uploads
  each one to the admin panel and marks the result.

  This way scraping and uploading happen in parallel — no waiting for the
  scraper to finish before uploading starts.

USAGE:
  Open TWO terminal windows inside HospitalReviewAutomation/.

  Terminal 1:   python scraper_main.py
  Terminal 2:   python upload_live.py

  Stop upload_live.py at any time with Ctrl+C.
  After scraper_main.py finishes, let upload_live.py drain remaining rows,
  then Ctrl+C it.

SAFETY:
  - Rows are claimed (set to "Pending") before uploading — no double-upload
    even if the script is restarted.
  - File access errors (scraper has the file locked) are retried silently.
  - All existing Upload Status values ("Uploaded", "Not Found on Site",
    "Failed", "Pending") are skipped — only blank rows are processed.
"""

import sys
import time
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout

try:
    from config import BASE_URL, USERNAME, PASSWORD, HEADLESS
    from logger import get_logger
    log = get_logger("UPLOAD_LIVE")
except ImportError:
    # Fallback if run outside the project folder
    BASE_URL  = "https://karunahealthlifepartner.com/index.php/account/admin"
    USERNAME  = "karuna__admin"
    PASSWORD  = "karuna__admin"
    HEADLESS  = True
    logging.basicConfig(
        level=logging.INFO,
        format="[%(asctime)s] %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[logging.StreamHandler(sys.stdout)],
    )
    log = logging.getLogger("UPLOAD_LIVE")

# =============================================================================
# CONFIG
# =============================================================================
EXCEL_FILE     = Path(__file__).resolve().parent / "output" / "HospitalAutomation.xlsx"
POLL_INTERVAL  = 30    # seconds between Excel scans
FILE_RETRY     = 5     # seconds to wait when Excel is locked by the scraper
FILE_ATTEMPTS  = 6     # how many times to retry a locked file before skipping

# Statuses that mean "already handled — skip"
SKIP_STATUSES = {"uploaded", "not found on site", "failed", "pending"}

# =============================================================================
# EXCEL HELPERS
# =============================================================================

def _load_wb_safe():
    """Open the workbook with retry in case the scraper has it locked."""
    for attempt in range(1, FILE_ATTEMPTS + 1):
        try:
            return load_workbook(EXCEL_FILE)
        except Exception as e:
            if attempt < FILE_ATTEMPTS:
                log.debug(f"Excel locked ({e}), retrying in {FILE_RETRY}s…")
                time.sleep(FILE_RETRY)
            else:
                raise


def claim_new_rows() -> list[dict]:
    """
    Find all rows in the Reviews sheet with a blank Upload Status,
    mark them as 'Pending' immediately (to prevent double-processing),
    and return their data.
    """
    try:
        wb = _load_wb_safe()
    except Exception as e:
        log.warning(f"Could not open Excel to scan: {e}")
        return []

    ws    = wb["Reviews"]
    found = []

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        status_cell = row[6]  # column G = Upload Status
        status_val  = str(status_cell.value or "").strip().lower()

        if status_val in SKIP_STATUSES:
            continue

        hospital_name = str(row[0].value or "").strip()
        review_text   = str(row[3].value or "").strip()

        if not hospital_name or not review_text:
            continue

        # Claim the row right now so another instance doesn't pick it up
        status_cell.value = "Pending"

        found.append({
            "hospital_name": hospital_name,
            "reviewer":      str(row[1].value or "Unknown").strip(),
            "rating":        row[2].value or 0,
            "review":        review_text,
            "review_date":   str(row[4].value or "").strip(),
            "source":        str(row[5].value or "").strip(),
            "row_index":     i,
        })

    if found:
        try:
            wb.save(EXCEL_FILE)
            log.info(f"Claimed {len(found)} new row(s) for upload.")
        except Exception as e:
            log.warning(f"Could not save claimed rows: {e}")
            wb.close()
            return []

    wb.close()
    return found


def mark_row(row_index: int, status: str) -> None:
    """Write the final Upload Status back to a specific row."""
    for attempt in range(1, FILE_ATTEMPTS + 1):
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Reviews"]
            ws.cell(row=row_index, column=7, value=status)
            wb.save(EXCEL_FILE)
            wb.close()
            return
        except Exception as e:
            if attempt < FILE_ATTEMPTS:
                log.debug(f"Retry marking row {row_index}: {e}")
                time.sleep(FILE_RETRY)
            else:
                log.error(f"Failed to mark row {row_index} as '{status}': {e}")

# =============================================================================
# BROWSER / UPLOAD HELPERS  (copied exactly from upload_reviews.py)
# =============================================================================

def login(page):
    log.info("Logging into admin panel…")
    page.goto(BASE_URL, wait_until="domcontentloaded", timeout=60_000)
    page.get_by_role("textbox", name="Enter User Name").fill(USERNAME)
    page.get_by_role("textbox", name="Password").fill(PASSWORD)
    page.get_by_role("button",  name="SIGN IN").click()
    page.wait_for_load_state("networkidle", timeout=30_000)
    log.info("Logged in.")


def go_to_reviews_page(page):
    page.get_by_role("link", name="Reviews & Ratings").click()
    page.wait_for_load_state("domcontentloaded", timeout=30_000)
    page.wait_for_timeout(1_000)


def open_add_review_form(page):
    page.get_by_role("listitem").filter(has_text="Add Review").click()
    page.wait_for_timeout(1_200)


def search_and_select_hospital(page, hospital_name: str) -> bool:
    search_box  = page.get_by_role("textbox", name="Type to search listings (e.g")
    search_term = hospital_name[:40].strip()
    search_box.fill("")
    search_box.type(search_term, delay=40)
    page.wait_for_timeout(3_000)

    options = page.locator("text=/ — /").all()
    if not options:
        first_word = hospital_name.split()[0] if hospital_name.split() else hospital_name
        if first_word.lower() != search_term.lower():
            search_box.fill("")
            search_box.type(first_word, delay=40)
            page.wait_for_timeout(3_000)
            options = page.locator("text=/ — /").all()

    if not options:
        log.warning(f"  No dropdown for: {hospital_name}")
        return False

    hospital_words = set(hospital_name.lower().split())
    best_score, best_option = -1, None
    for opt in options:
        try:
            text = (opt.inner_text() or "").strip()
        except Exception:
            continue
        if not text or "no results" in text.lower():
            continue
        score = len(hospital_words & set(text.lower().split()))
        if score > best_score:
            best_score, best_option = score, opt

    if best_option is None or best_score == 0:
        log.warning(f"  No match for: {hospital_name}")
        return False

    try:
        best_option.click()
        page.wait_for_timeout(600)
        return True
    except Exception as e:
        log.warning(f"  Click failed: {e}")
        return False


def click_star(page, rating):
    try:
        n = int(float(str(rating)))
    except (ValueError, TypeError):
        n = 0
    if n < 1 or n > 5:
        return
    try:
        page.locator(f"#amrStarRow > i:nth-child({n})").click()
        page.wait_for_timeout(400)
    except Exception as e:
        log.warning(f"  Star click failed: {e}")


def upload_one_review(page, review: dict) -> str:
    """Returns 'Uploaded', 'Not Found on Site', or 'Failed'."""
    try:
        page.get_by_role("textbox", name="e.g. Rahul S., Priya G.,").fill(review["reviewer"])
        if not search_and_select_hospital(page, review["hospital_name"]):
            return "Not Found on Site"
        click_star(page, review["rating"])
        text = review["review"] or "Good experience at this hospital."
        page.get_by_role("textbox", name="Write the review content…").fill(text)
        page.get_by_role("radio", name="Pending (need to approve").check()
        page.wait_for_timeout(300)
        page.get_by_role("button", name="Save Review").click()
        page.wait_for_timeout(2_000)
        return "Uploaded"
    except Exception as e:
        log.error(f"  Upload error for {review['hospital_name']}: {e}")
        return "Failed"


def safe_go_to_reviews(page):
    """Navigate back to the Reviews page; re-login if the session expired."""
    try:
        go_to_reviews_page(page)
    except Exception:
        log.warning("Navigation failed — re-logging in.")
        try:
            login(page)
            go_to_reviews_page(page)
        except Exception as e:
            log.error(f"Re-login failed: {e}")
            raise

# =============================================================================
# MAIN WATCH LOOP
# =============================================================================

def main():
    print("=" * 70)
    print("  LIVE UPLOAD WATCHER — upload_live.py")
    print(f"  Excel file    : {EXCEL_FILE}")
    print(f"  Poll interval : every {POLL_INTERVAL}s")
    print(f"  Admin URL     : {BASE_URL}")
    print(f"  Started at    : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    print("\nRun scraper_main.py in another terminal window.")
    print("This script will upload reviews as they are scraped.")
    print("Press Ctrl+C to stop.\n")

    if not EXCEL_FILE.exists():
        sys.exit(f"[FATAL] Excel file not found: {EXCEL_FILE}")

    uploaded_total  = 0
    not_found_total = 0
    failed_total    = 0

    with sync_playwright() as pw:
        browser = pw.chromium.launch(headless=HEADLESS)
        page    = browser.new_context(
            locale="en-US",
            viewport={"width": 1280, "height": 900},
        ).new_page()

        try:
            login(page)
            go_to_reviews_page(page)

            while True:
                # ── Scan Excel for newly scraped rows ──────────────────────
                new_rows = claim_new_rows()

                if not new_rows:
                    print(f"[{datetime.now().strftime('%H:%M:%S')}] "
                          f"No new reviews yet. "
                          f"Uploaded so far: {uploaded_total} ✓ | "
                          f"Not found: {not_found_total} | "
                          f"Failed: {failed_total}. "
                          f"Checking again in {POLL_INTERVAL}s…")
                    time.sleep(POLL_INTERVAL)
                    continue

                print(f"\n[{datetime.now().strftime('%H:%M:%S')}] "
                      f"Found {len(new_rows)} new review(s) — uploading now…")

                # ── Upload each claimed row ────────────────────────────────
                for idx, review in enumerate(new_rows, start=1):
                    print(f"  [{idx}/{len(new_rows)}] "
                          f"{review['hospital_name']} | "
                          f"{review['reviewer']} | "
                          f"{review['rating']}★")

                    result = "Failed"
                    try:
                        open_add_review_form(page)
                        result = upload_one_review(page, review)
                    except Exception as e:
                        log.error(f"Unhandled error on row {review['row_index']}: {e}")
                        result = "Failed"

                    mark_row(review["row_index"], result)

                    if result == "Uploaded":
                        uploaded_total += 1
                        print(f"     ✓ Uploaded  (total: {uploaded_total})")
                    elif result == "Not Found on Site":
                        not_found_total += 1
                        print(f"     ✗ Not Found on Site")
                    else:
                        failed_total += 1
                        print(f"     ✗ Failed")

                    # Navigate back for next review (re-login if session dropped)
                    try:
                        safe_go_to_reviews(page)
                    except Exception:
                        log.error("Could not return to Reviews page — stopping.")
                        raise

                # ── Wait before next scan ──────────────────────────────────
                print(f"\n  Batch done. Waiting {POLL_INTERVAL}s for more…")
                time.sleep(POLL_INTERVAL)

        except KeyboardInterrupt:
            print("\n\n[STOPPED] Ctrl+C received.")

        finally:
            browser.close()

    print("\n" + "=" * 70)
    print("  LIVE UPLOAD COMPLETE")
    print(f"  Uploaded        : {uploaded_total}")
    print(f"  Not Found       : {not_found_total}")
    print(f"  Failed          : {failed_total}")
    print(f"  Finished at     : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)


if __name__ == "__main__":
    main()
