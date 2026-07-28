"""
collect_reviews.py
------------------
Phase 2 — Review Collection

Reads all hospitals from HospitalAutomation.xlsx,
finds their review pages on HexaHealth / Practo / Justdial,
scrapes up to 5 reviews each, writes results to Excel,
and saves a checkpoint after every hospital so the run
can be resumed after a crash.

Run:
    python collect_reviews.py

Resume after crash:
    Just run again — it picks up from the checkpoint automatically.
"""

import re
import requests
from pathlib import Path
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

from config import (
    HEADLESS,
    MAX_REVIEWS_PER_HOSPITAL,
    LOW_REVIEW_THRESHOLD,
)
from logger import get_logger
from checkpoint import save_checkpoint, load_checkpoint, clear_checkpoint
from excel_writer import ExcelWriter
from review_finder import find_review_url
from review_scraper import scrape_reviews   # HexaHealth scraper (untouched)

log = get_logger("COLLECT_REVIEWS")

EXCEL_FILE = Path("output/HospitalAutomation.xlsx")

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/138.0 Safari/537.36"
    )
}

# Set to an integer to limit hospitals during testing; None = all.
MAX_HOSPITALS = None


# ─────────────────────────────────────────────
# Site-specific scrapers
# ─────────────────────────────────────────────

def _scrape_practo(url: str, max_reviews: int = 5, page=None) -> list[dict]:
    """
    Scrape reviews from a Practo hospital page.
    Returns list of {reviewer, rating, review, date}.

    When `page` (a Playwright Page) is supplied, the already-rendered HTML is
    used directly — avoids 404s on JS-rendered pages.
    """
    if page is not None:
        html = page.content()
    else:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        html = resp.text
    soup = BeautifulSoup(html, "html.parser")

    reviews = []

    # Practo uses several different layouts — try each selector in order.
    cards = (
        soup.select("div.doctor-review-text")
        or soup.select("div[class*='review-card']")
        or soup.select("div[class*='review_card']")
        or soup.select("div[class*='ReviewCard']")
        or soup.select("div[data-qa-id='review_card']")
    )

    for card in cards:
        # Reviewer name
        name_tag = (
            card.find("p",  class_=re.compile(r"reviewer|author|username", re.I))
            or card.find("span", class_=re.compile(r"reviewer|author|username", re.I))
            or card.find("strong")
        )
        reviewer = name_tag.get_text(strip=True) if name_tag else "Unknown"

        # Rating — count filled stars or read numeric value
        stars = card.select("i.fa-star, span.fa-star, img[alt*='star']")
        rating = len(stars) if stars else 0
        if rating == 0:
            rating_tag = card.find(class_=re.compile(r"rating|score", re.I))
            if rating_tag:
                try:
                    rating = float(re.search(r"\d+\.?\d*", rating_tag.get_text()).group())
                except Exception:
                    rating = 0

        # Review text
        text_tag = (
            card.find("p",   class_=re.compile(r"review.?text|comment|body", re.I))
            or card.find("div", class_=re.compile(r"review.?text|comment|body", re.I))
            or card.find("p")
        )
        review_text = text_tag.get_text(strip=True) if text_tag else ""

        # Date
        date_tag = card.find(class_=re.compile(r"date|time|when", re.I))
        date = date_tag.get_text(strip=True) if date_tag else ""

        reviews.append({
            "reviewer": reviewer,
            "rating":   rating,
            "review":   review_text,
            "date":     date,
        })

        if len(reviews) >= max_reviews:
            break

    return reviews


def _scrape_justdial(url: str, max_reviews: int = 5, page=None) -> list[dict]:
    """
    Scrape reviews from a Justdial listing page.
    Returns list of {reviewer, rating, review, date}.

    When `page` (a Playwright Page) is supplied, the already-rendered HTML is
    used directly — avoids 404s on JS-rendered pages.
    """
    if page is not None:
        html = page.content()
    else:
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        html = resp.text
    soup = BeautifulSoup(html, "html.parser")

    reviews = []

    cards = (
        soup.select("div.review-box")
        or soup.select("div[class*='reviewdesc']")
        or soup.select("div[class*='review_box']")
        or soup.select("div[class*='ReviewBox']")
        or soup.select("li.review-item")
    )

    for card in cards:
        # Reviewer name
        name_tag = (
            card.find(class_=re.compile(r"reviewer|user.?name|author", re.I))
            or card.find("span")
        )
        reviewer = name_tag.get_text(strip=True) if name_tag else "Unknown"

        # Rating
        stars = card.select("i.icon-star-full, span[class*='star'], em[class*='star']")
        rating = len(stars) if stars else 0
        if rating == 0:
            rating_tag = card.find(class_=re.compile(r"rating|score|star", re.I))
            if rating_tag:
                try:
                    rating = float(re.search(r"\d+\.?\d*", rating_tag.get_text()).group())
                except Exception:
                    rating = 0

        # Review text
        text_tag = (
            card.find("p",   class_=re.compile(r"review.?text|comment|desc", re.I))
            or card.find("div", class_=re.compile(r"review.?text|comment|desc", re.I))
            or card.find("p")
        )
        review_text = text_tag.get_text(strip=True) if text_tag else ""

        # Date
        date_tag = card.find(class_=re.compile(r"date|time|posted", re.I))
        date = date_tag.get_text(strip=True) if date_tag else ""

        reviews.append({
            "reviewer": reviewer,
            "rating":   rating,
            "review":   review_text,
            "date":     date,
        })

        if len(reviews) >= max_reviews:
            break

    return reviews


def scrape_for_site(site: str, url: str, max_reviews: int = 5, page=None) -> list[dict]:
    """
    Dispatch to the right scraper based on which site the URL is from.
    Falls back to the HexaHealth scraper if site is unrecognised.

    Pass `page` (Playwright Page) so scrapers can use the already-rendered
    browser HTML instead of a plain HTTP request — required for SPAs that
    return 404 to requests but render fine in a real browser.
    """
    if site == "hexahealth.com":
        return scrape_reviews(url, max_reviews=max_reviews, page=page)
    elif site == "practo.com":
        return _scrape_practo(url, max_reviews=max_reviews, page=page)
    elif site == "justdial.com":
        return _scrape_justdial(url, max_reviews=max_reviews, page=page)
    else:
        # Generic fallback
        return scrape_reviews(url, max_reviews=max_reviews, page=page)


# ─────────────────────────────────────────────
# Load hospitals
# ─────────────────────────────────────────────

def load_hospitals() -> list[dict]:
    """Read all rows from the Hospitals sheet."""
    if not EXCEL_FILE.exists():
        raise FileNotFoundError(
            f"Excel file not found: {EXCEL_FILE}\n"
            "Run extract_hospitals.py first."
        )

    wb = load_workbook(EXCEL_FILE)
    ws = wb["Hospitals"]

    hospitals = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[1]:
            continue
        hospitals.append({
            "id":       row[0],
            "name":     row[1],
            "phone":    row[2],
            "email":    row[3],
            "status":   row[4],
            "edit_url": row[5],
        })

    wb.close()
    log.info(f"Loaded {len(hospitals)} hospitals from Excel.")
    return hospitals


# ─────────────────────────────────────────────
# Process one hospital
# ─────────────────────────────────────────────

def process_hospital(page, hospital: dict, writer: ExcelWriter) -> str:
    """
    Find and scrape reviews for one hospital, then write to Excel.

    Returns the collection status:
      "Success" | "Low Reviews" | "No Reviews" | "No Match" | "Scraper Error"
    """
    name = hospital["name"]
    log.info(f"Processing: {name}")
    print(f"\n{'='*70}")
    print(f"Hospital : {name}")

    # ── Step 1: Find review URL ──────────────────────────────────────────
    try:
        site, url = find_review_url(page, name)
    except Exception as exc:
        log.error(f"review_finder crashed for {name}: {exc}")
        site, url = None, None

    if not url:
        print("  Source   : None")
        print("  Status   : No Match")
        writer.add_summary_row(
            hospital=name,
            status="No Match",
            reviews_found=0,
            reason="Not found on HexaHealth, Practo, or Justdial",
        )
        writer.update_collection_status(name, 0, "No Match")
        return "No Match"

    print(f"  Source   : {site}")
    print(f"  URL      : {url}")

    # ── Step 2: Scrape reviews using the right scraper ───────────────────
    try:
        reviews = scrape_for_site(site, url, max_reviews=MAX_REVIEWS_PER_HOSPITAL, page=page)
    except Exception as exc:
        log.error(f"Scraper failed for {name} ({site}): {exc}")
        print(f"  Status   : Scraper Error — {exc}")
        writer.add_summary_row(
            hospital=name,
            status="Scraper Error",
            reviews_found=0,
            reason=str(exc),
            source=site,
            url=url,
        )
        writer.update_collection_status(name, 0, "Scraper Error")
        return "Scraper Error"

    count = len(reviews)
    print(f"  Reviews  : {count}")

    # ── Step 3: Categorise ───────────────────────────────────────────────
    if count == 0:
        status = "No Reviews"
        reason = "Review page found but contains no scrapeable reviews"
    elif count < LOW_REVIEW_THRESHOLD:
        status = "Low Reviews"
        reason = f"Only {count} review(s) found (threshold: {LOW_REVIEW_THRESHOLD})"
    else:
        status = "Success"
        reason = ""

    # ── Step 4: Write reviews to Excel ───────────────────────────────────
    for r in reviews:
        writer.add_review(
            hospital_name=name,
            reviewer=r.get("reviewer", "Unknown"),
            rating=r.get("rating", ""),
            review=r.get("review", ""),
            review_date=r.get("date", ""),
            source=site,
            upload_status="Pending",
        )
        print(f"  → {r.get('reviewer','Unknown')} | {r.get('rating','?')}★")

    # ── Step 5: Write summary row ────────────────────────────────────────
    writer.add_summary_row(
        hospital=name,
        status=status,
        reviews_found=count,
        reason=reason,
        source=site,
        url=url,
    )

    # ── Step 6: Update Hospitals sheet ───────────────────────────────────
    writer.update_collection_status(name, count, status)

    print(f"  Status   : {status}")
    return status


# ─────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────

def main():
    hospitals = load_hospitals()
    total = len(hospitals)

    if MAX_HOSPITALS is not None:
        hospitals = hospitals[:MAX_HOSPITALS]
        print(f"TEST MODE: limiting to {MAX_HOSPITALS} hospitals.")

    # Resume from checkpoint if a previous run was interrupted
    checkpoint = load_checkpoint()
    if checkpoint and checkpoint.get("phase") == "collect":
        start_index = checkpoint["hospital_index"]
        print(f"\nResuming from hospital #{start_index}: {checkpoint['hospital_name']}")
        log.info(f"Resuming collect from index {start_index}")
    else:
        start_index = 0
        print(f"\nStarting fresh collection for {total} hospitals.")

    writer = ExcelWriter()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        page = browser.new_page()

        try:
            for i, hospital in enumerate(hospitals[start_index:], start=start_index):
                try:
                    process_hospital(page, hospital, writer)
                except Exception as exc:
                    log.error(f"Unhandled error for {hospital['name']}: {exc}")
                    print(f"  ERROR: {exc} — skipping.")

                # Save progress after every single hospital
                save_checkpoint(
                    phase="collect",
                    hospital_index=i + 1,
                    hospital_name=hospital["name"],
                )

                print(f"  Progress : {i + 1}/{total}")

        finally:
            browser.close()

    clear_checkpoint()
    print("\n" + "="*70)
    print("Collection complete.")
    print(f"Results saved to: {EXCEL_FILE}")


if __name__ == "__main__":
    main()
