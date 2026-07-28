"""
excel_writer.py
---------------
Excel helper for the Hospital Review Automation project.

Manages HospitalAutomation.xlsx with four sheets:
  - Hospitals     (written by extract_hospitals.py)
  - Reviews       (written by collect_reviews.py)
  - Upload Status (written by upload_reviews.py)
  - Summary       (written by collect_reviews.py)

DO NOT remove or rename existing methods — other modules depend on them.
"""

from openpyxl import Workbook, load_workbook
from pathlib import Path

from logger import get_logger

log = get_logger("EXCEL")


class ExcelWriter:

    def __init__(self):

        self.file = Path("output/HospitalAutomation.xlsx")

        if not self.file.exists():

            wb = Workbook()

            hospitals = wb.active
            hospitals.title = "Hospitals"

            hospitals.append([
                "Internal ID",
                "Hospital Name",
                "Phone",
                "Email",
                "Status",
                "Edit URL",
                "Google Maps URL",
                "Reviews Found",
                "Collection Status",
            ])

            reviews = wb.create_sheet("Reviews")
            reviews.append([
                "Hospital Name",
                "Reviewer",
                "Rating",
                "Review",
                "Review Date",
                "Source",
                "Upload Status",
            ])

            upload = wb.create_sheet("Upload Status")
            upload.append([
                "Hospital Name",
                "Reviews Uploaded",
                "Upload Date",
                "Status",
                "Notes",
            ])

            summary = wb.create_sheet("Summary")
            summary.append([
                "Hospital",
                "Status",
                "Reviews Found",
                "Reason",
                "Source",
                "URL",
            ])

            wb.save(self.file)
            log.info("Created HospitalAutomation.xlsx")

        else:
            # Ensure all sheets + headers exist in an existing workbook.
            self._ensure_sheets()

    # ─────────────────────────────────────────
    # Internal
    # ─────────────────────────────────────────

    def _ensure_sheets(self):
        """Add any missing sheets with their headers."""
        wb = load_workbook(self.file)
        changed = False

        sheet_headers = {
            "Reviews": [
                "Hospital Name", "Reviewer", "Rating",
                "Review", "Review Date", "Source", "Upload Status",
            ],
            "Upload Status": [
                "Hospital Name", "Reviews Uploaded",
                "Upload Date", "Status", "Notes",
            ],
            "Summary": [
                "Hospital", "Status", "Reviews Found",
                "Reason", "Source", "URL",
            ],
        }

        for name, headers in sheet_headers.items():
            if name not in wb.sheetnames:
                ws = wb.create_sheet(name)
                ws.append(headers)
                changed = True
            else:
                ws = wb[name]
                if ws.max_row == 0 or all(c.value is None for c in ws[1]):
                    ws.insert_rows(1)
                    for col, h in enumerate(headers, start=1):
                        ws.cell(row=1, column=col, value=h)
                    changed = True

        if changed:
            wb.save(self.file)
        wb.close()

    # ─────────────────────────────────────────
    # Hospitals sheet (unchanged from original)
    # ─────────────────────────────────────────

    def add_hospital(self, hospital: dict) -> bool:
        """
        Append a hospital row if it doesn't already exist.
        Returns True if added, False if skipped (duplicate).
        """
        wb = load_workbook(self.file)
        ws = wb["Hospitals"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == str(hospital["id"]):
                wb.close()
                return False

        ws.append([
            hospital["id"],
            hospital["name"],
            hospital.get("phone", ""),
            hospital.get("email", ""),
            hospital.get("status", ""),
            hospital.get("edit_url", ""),
            "",
            "",
            "Pending",
        ])

        wb.save(self.file)
        wb.close()
        log.info(f"Added hospital: {hospital['name']}")
        return True

    # ─────────────────────────────────────────
    # Reviews sheet
    # ─────────────────────────────────────────

    def add_review(
        self,
        hospital_name: str,
        reviewer: str,
        rating,
        review: str,
        review_date: str,
        source: str,
        upload_status: str = "Pending",
    ):
        """Append one review row to the Reviews sheet."""
        wb = load_workbook(self.file)
        ws = wb["Reviews"]

        ws.append([
            hospital_name,
            reviewer,
            rating,
            review,
            review_date,
            source,
            upload_status,
        ])

        wb.save(self.file)
        wb.close()
        log.info(f"Added review: {hospital_name} — {reviewer}")

    def update_upload_status_in_reviews(self, hospital_name: str, new_status: str):
        """
        Update the Upload Status column for every review row
        belonging to `hospital_name`.
        """
        wb = load_workbook(self.file)
        ws = wb["Reviews"]

        for row in ws.iter_rows(min_row=2):
            if row[0].value == hospital_name:
                row[6].value = new_status

        wb.save(self.file)
        wb.close()
        log.info(f"Updated upload status for {hospital_name} → {new_status}")

    # ─────────────────────────────────────────
    # Summary sheet
    # ─────────────────────────────────────────

    def add_summary_row(
        self,
        hospital: str,
        status: str,
        reviews_found: int,
        reason: str = "",
        source: str = "",
        url: str = "",
    ):
        """Append one row to the Summary sheet."""
        wb = load_workbook(self.file)
        ws = wb["Summary"]

        ws.append([hospital, status, reviews_found, reason, source, url])

        wb.save(self.file)
        wb.close()
        log.info(f"Summary row: {hospital} → {status} ({reviews_found} reviews)")

    # ─────────────────────────────────────────
    # Upload Status sheet
    # ─────────────────────────────────────────

    def add_upload_status_row(
        self,
        hospital_name: str,
        reviews_uploaded: int,
        upload_date: str,
        status: str,
        notes: str = "",
    ):
        """Append one row to the Upload Status sheet."""
        wb = load_workbook(self.file)
        ws = wb["Upload Status"]

        ws.append([
            hospital_name,
            reviews_uploaded,
            upload_date,
            status,
            notes,
        ])

        wb.save(self.file)
        wb.close()
        log.info(f"Upload status recorded: {hospital_name} → {status}")

    # ─────────────────────────────────────────
    # Hospitals sheet — collection status update
    # ─────────────────────────────────────────

    def update_collection_status(
        self,
        hospital_name: str,
        reviews_found: int,
        collection_status: str,
    ):
        """
        Update Reviews Found (col H) and Collection Status (col I)
        in the Hospitals sheet for the given hospital.
        """
        wb = load_workbook(self.file)
        ws = wb["Hospitals"]

        for row in ws.iter_rows(min_row=2):
            if row[1].value == hospital_name:
                row[7].value = reviews_found
                row[8].value = collection_status
                break

        wb.save(self.file)
        wb.close()
