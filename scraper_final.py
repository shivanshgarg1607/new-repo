"""
scraper_final.py  v13.0
------------------------
Unified scraper that runs BOTH phases (search + collect) using the proven
review_finder.py (DuckDuckGo in a real browser) and review_scraper.py
modules — the same approach that collected ~750 reviews in collect_reviews.py.

Key fixes vs old scraper_final.py:
  1. Uses review_finder.py (DDG in real browser) instead of Google HTTP.
  2. Browser is VISIBLE (headless=False) so you can watch it work.
  3. Does NOT block images/fonts — full page rendering for reliable scraping.
  4. Uses review_scraper.py + site-specific parsers from collect_reviews.py.
  5. Two-phase approach retained but both phases use the proven search method.

Phase 1 — hospitals whose Collection Status is not "Success".
Phase 2 — gap fill on already-successful hospitals to reach the target.

Pre-flight (run once):
    pip install playwright openpyxl beautifulsoup4 requests lxml
    python -m playwright install chromium
"""

import re
import sys
import json
import time
import random
import hashlib
import logging
import traceback
from datetime import datetime
from pathlib import Path
from urllib.parse import quote_plus

try:
    from bs4 import BeautifulSoup
    from openpyxl import load_workbook
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError as _e:
    sys.exit(
        f"\n[FATAL] Missing dependency: {_e}\n"
        "Run:\n"
        "  pip install playwright openpyxl beautifulsoup4 requests lxml\n"
        "  python -m playwright install chromium\n"
    )

from config import (
    HEADLESS,
    MAX_REVIEWS_PER_HOSPITAL,
    LOW_REVIEW_THRESHOLD,
    SEARCH_DELAY_MIN,
    SEARCH_DELAY_MAX,
)
from logger import get_logger
from checkpoint import save_checkpoint, load_checkpoint, clear_checkpoint
from excel_writer import ExcelWriter
from review_finder import find_review_url
from review_scraper import scrape_reviews

# ─────────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────────

_ROOT           = Path(__file__).resolve().parent
EXCEL_FILE      = _ROOT / "output" / "HospitalAutomation.xlsx"
LOG_DIR         = _ROOT / "logs"
CACHE_DIR       = _ROOT / "cache"
LOG_DIR.mkdir(exist_ok=True)
CACHE_DIR.mkdir(exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# Config
# ─────────────────────────────────────────────────────────────────────────────

TARGET_NEW_REVIEWS      = 1500
PHASE1_MAX_PER_HOSPITAL = 50      # max reviews per hospital in phase 1
PHASE2_PER_HOSPITAL     = 10      # extra reviews per successful hospital in phase 2

BROWSER_TIMEOUT    = 30_000
PAGE_WAIT_MS       = 2_500
SCROLL_WAIT_MS     = 1_500
SCROLL_ROUNDS      = 20

# Hospitals whose Collection Status is one of these need fresh searching.
NON_SUCCESS_STATUSES = {
    "No Match", "No Reviews", "Failed", "Scraper Error",
    "No Review", "Not Found", "Error", "Pending", "Low Reviews", None, "",
}

# ─────────────────────────────────────────────────────────────────────────────
# Logging
# ─────────────────────────────────────────────────────────────────────────────

_log_file = LOG_DIR / f"scraper_final_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(_log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("scraper_final")

# ─────────────────────────────────────────────────────────────────────────────
# Site-specific scrapers (from collect_reviews.py — proven to work)
# ─────────────────────────────────────────────────────────────────────────────

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/138.0 Safari/537.36"
    )
}


def _scrape_practo(url: str, max_reviews: int = 50, page=None) -> list[dict]:
    """Scrape reviews from a Practo hospital page using rendered browser HTML."""
    if page is not None:
        prev_count = 0
        for _scroll in range(SCROLL_ROUNDS):
            try:
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                page.wait_for_timeout(2_000)
            except Exception:
                pass
            try:
                cur_count = page.locator(
                    "div.doctor-review-text, div[class*='review-card'], "
                    "div[class*='review_card'], div[class*='ReviewCard'], "
                    "div[data-qa-id='review_card']"
                ).count()
            except Exception:
                cur_count = 0
            if cur_count <= prev_count and _scroll >= 2:
                break
            prev_count = cur_count
        html = page.content()
    else:
        import requests
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        html = resp.text

    soup = BeautifulSoup(html, "html.parser")
    reviews = []
    cards = (
        soup.select("div.doctor-review-text")
        or soup.select("div[class*='review-card']")
        or soup.select("div[class*='review_card']")
        or soup.select("div[class*='ReviewCard']")
        or soup.select("div[data-qa-id='review_card']")
    )

    for card in cards:
        name_tag = (
            card.find("p", class_=re.compile(r"reviewer|author|username", re.I))
            or card.find("span", class_=re.compile(r"reviewer|author|username", re.I))
            or card.find("strong")
        )
        reviewer = name_tag.get_text(strip=True) if name_tag else "Unknown"

        stars = card.select("i.fa-star, span.fa-star, img[alt*='star']")
        rating = len(stars) if stars else 0
        if rating == 0:
            rating_tag = card.find(class_=re.compile(r"rating|score", re.I))
            if rating_tag:
                try:
                    rating = float(re.search(r"\d+\.?\d*", rating_tag.get_text()).group())
                except Exception:
                    rating = 0

        text_tag = (
            card.find("p", class_=re.compile(r"review.?text|comment|body", re.I))
            or card.find("div", class_=re.compile(r"review.?text|comment|body", re.I))
            or card.find("p")
        )
        review_text = text_tag.get_text(strip=True) if text_tag else ""

        date_tag = card.find(class_=re.compile(r"date|time|when", re.I))
        date = date_tag.get_text(strip=True) if date_tag else ""

        reviews.append({
            "reviewer": reviewer,
            "rating": rating,
            "review": review_text,
            "date": date,
        })
        if len(reviews) >= max_reviews:
            break
    return reviews


def _scrape_justdial(url: str, max_reviews: int = 50, page=None) -> list[dict]:
    """Scrape reviews from a Justdial listing page using rendered browser HTML."""
    if page is not None:
        prev_count = 0
        for _scroll in range(SCROLL_ROUNDS):
            try:
                page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
                page.wait_for_timeout(2_000)
            except Exception:
                pass
            try:
                cur_count = page.locator(
                    "div.review-box, div[class*='reviewdesc'], "
                    "div[class*='review_box'], div[class*='ReviewBox'], "
                    "li.review-item"
                ).count()
            except Exception:
                cur_count = 0
            if cur_count <= prev_count and _scroll >= 2:
                break
            prev_count = cur_count
        html = page.content()
    else:
        import requests
        resp = requests.get(url, headers=HEADERS, timeout=30)
        resp.raise_for_status()
        html = resp.text

    soup = BeautifulSoup(html, "html.parser")
    reviews = []
    cards = (
        soup.select("div.review-box")
        or soup.select("div[class*='reviewdesc']")
        or soup.select("div[class*='review_box']")
        or soup.select("div[class*='ReviewBox']")
        or soup.select("li.review-item")
    )

    for card in cards:
        name_tag = (
            card.find(class_=re.compile(r"reviewer|user.?name|author", re.I))
            or card.find("span")
        )
        reviewer = name_tag.get_text(strip=True) if name_tag else "Unknown"

        stars = card.select("i.icon-star-full, span[class*='star'], em[class*='star']")
        rating = len(stars) if stars else 0
        if rating == 0:
            rating_tag = card.find(class_=re.compile(r"rating|score|star", re.I))
            if rating_tag:
                try:
                    rating = float(re.search(r"\d+\.?\d*", rating_tag.get_text()).group())
                except Exception:
                    rating = 0

        text_tag = (
            card.find("p", class_=re.compile(r"review.?text|comment|desc", re.I))
            or card.find("div", class_=re.compile(r"review.?text|comment|desc", re.I))
            or card.find("p")
        )
        review_text = text_tag.get_text(strip=True) if text_tag else ""

        date_tag = card.find(class_=re.compile(r"date|time|posted", re.I))
        date = date_tag.get_text(strip=True) if date_tag else ""

        reviews.append({
            "reviewer": reviewer,
            "rating": rating,
            "review": review_text,
            "date": date,
        })
        if len(reviews) >= max_reviews:
            break
    return reviews


def scrape_for_site(site: str, url: str, max_reviews: int = 50, page=None) -> list[dict]:
    """Dispatch to the right scraper based on which site the URL is from."""
    if site == "hexahealth.com":
        return scrape_reviews(url, max_reviews=max_reviews, page=page)
    elif site == "practo.com":
        return _scrape_practo(url, max_reviews=max_reviews, page=page)
    elif site == "justdial.com":
        return _scrape_justdial(url, max_reviews=max_reviews, page=page)
    else:
        return scrape_reviews(url, max_reviews=max_reviews, page=page)


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

def load_hospitals() -> list[dict]:
    """
    Read all hospital names + collection statuses from the Excel file.
    Prefers the 'Hospitals' sheet; falls back to the 'Summary' sheet if
    'Hospitals' is not present (some copies of the workbook only have
    Reviews + Summary).
    """
    if not EXCEL_FILE.exists():
        sys.exit(f"\n[FATAL] Excel file not found: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE, read_only=True)

    out = []

    if "Hospitals" in wb.sheetnames:
        ws = wb["Hospitals"]
        for rn, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = str(row[1] or "").strip() if len(row) > 1 else ""
            if not name:
                continue
            collection_status = str(row[8] or "").strip() if len(row) > 8 else ""
            out.append({
                "name":              name,
                "collection_status": collection_status,
                "row_number":        rn,
            })
    elif "Summary" in wb.sheetnames:
        ws = wb["Summary"]
        for rn, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            name = str(row[0] or "").strip() if len(row) > 0 else ""
            if not name:
                continue
            status = str(row[1] or "").strip() if len(row) > 1 else ""
            out.append({
                "name":              name,
                "collection_status": status,
                "row_number":        rn,
            })
    else:
        sys.exit(
            f"\n[FATAL] No 'Hospitals' or 'Summary' sheet found. "
            f"Sheets: {wb.sheetnames}"
        )

    wb.close()
    log.info(f"Loaded {len(out):,} hospitals from Excel.")
    return out


def load_existing_review_keys() -> set:
    """Load existing review texts keyed by hospital name for dedup."""
    wb = load_workbook(EXCEL_FILE, read_only=True)
    ws = wb["Reviews"]
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        hospital = str(row[0] or "").strip().lower()
        text = str(row[3] or "").strip().lower()
        if hospital and text:
            keys.add(f"{hospital}|{text}")
    wb.close()
    log.info(f"Loaded {len(keys):,} existing review keys for dedup.")
    return keys


# ─────────────────────────────────────────────────────────────────────────────
# Process one hospital
# ─────────────────────────────────────────────────────────────────────────────

def process_hospital(page, hospital: dict, writer: ExcelWriter,
                      existing_keys: set, max_reviews: int) -> int:
    """
    Find and scrape reviews for one hospital, write to Excel.
    Returns the number of NEW reviews added.
    """
    name = hospital["name"]
    log.info(f"Processing: {name}")
    print(f"\n{'='*70}")
    print(f"Hospital : {name}")

    # ── Step 1: Find review URL using proven review_finder.py ─────────────
    try:
        site, url = find_review_url(page, name)
    except Exception as exc:
        log.error(f"review_finder crashed for {name}: {exc}")
        site, url = None, None

    if not url:
        print("  Source   : None")
        print("  Status   : No Match")
        writer.add_summary_row(
            hospital=name,
            status="No Match",
            reviews_found=0,
            reason="Not found on HexaHealth, Practo, or Justdial",
        )
        writer.update_collection_status(name, 0, "No Match")
        return 0

    print(f"  Source   : {site}")
    print(f"  URL      : {url}")

    # ── Step 2: Scrape reviews using the right scraper ────────────────────
    # Load existing review texts from Excel for this hospital to skip dups
    from openpyxl import load_workbook as _lwb
    hosp_existing = set()
    try:
        _wb = _lwb(writer.file)
        _ws = _wb["Reviews"]
        for _row in _ws.iter_rows(min_row=2, values_only=True):
            if _row[0] == name:
                _text = str(_row[3] or "").strip().lower()
                if _text:
                    hosp_existing.add(_text)
        _wb.close()
    except Exception as _exc:
        log.warning(f"Could not read existing reviews for {name}: {_exc}")

    try:
        fetch_target = max_reviews + len(hosp_existing)
        reviews = scrape_for_site(site, url, max_reviews=fetch_target, page=page)
    except Exception as exc:
        log.error(f"Scraper failed for {name} ({site}): {exc}")
        print(f"  Status   : Scraper Error — {exc}")
        writer.add_summary_row(
            hospital=name,
            status="Scraper Error",
            reviews_found=0,
            reason=str(exc),
            source=site,
            url=url,
        )
        writer.update_collection_status(name, 0, "Scraper Error")
        return 0

    # Drop reviews already stored from a previous run
    unique_reviews = []
    for r in reviews:
        key = (r.get("review", "") or "").strip().lower()
        if key and key in hosp_existing:
            continue
        hosp_existing.add(key)
        unique_reviews.append(r)
        if len(unique_reviews) >= max_reviews:
            break

    reviews = unique_reviews
    count = len(reviews)
    print(f"  Reviews  : {count}")

    # ── Step 3: Categorise ───────────────────────────────────────────────
    if count == 0:
        status = "No Reviews"
        reason = "Review page found but contains no scrapeable reviews"
    elif count < LOW_REVIEW_THRESHOLD:
        status = "Low Reviews"
        reason = f"Only {count} review(s) found (threshold: {LOW_REVIEW_THRESHOLD})"
    else:
        status = "Success"
        reason = ""

    # ── Step 4: Write reviews to Excel ───────────────────────────────────
    for r in reviews:
        writer.add_review(
            hospital_name=name,
            reviewer=r.get("reviewer", "Unknown"),
            rating=r.get("rating", ""),
            review=r.get("review", ""),
            review_date=r.get("date", ""),
            source=site,
            upload_status="",
        )
        print(f"  -> {r.get('reviewer','Unknown')} | {r.get('rating','?')} star")

    # ── Step 5: Write summary + update hospital status ────────────────────
    writer.add_summary_row(
        hospital=name,
        status=status,
        reviews_found=count,
        reason=reason,
        source=site,
        url=url,
    )
    writer.update_collection_status(name, count, status)

    print(f"  Status   : {status}")
    return count


# ─────────────────────────────────────────────────────────────────────────────
# Phase 1 — Non-success hospitals (need fresh search)
# ─────────────────────────────────────────────────────────────────────────────

def run_phase1(page, hospitals, writer, existing_keys, checkpoint):
    candidates = [h for h in hospitals if h["collection_status"] in NON_SUCCESS_STATUSES]
    total, total_new, start = len(candidates), 0, 0

    if checkpoint and checkpoint.get("phase") == "phase1":
        start = checkpoint.get("hospital_index", 0)
        total_new = checkpoint.get("total_new", total_new)
        print(f"\n[RESUME] Phase 1 from #{start} | Collected: {total_new:,}")

    print(f"\n{'='*70}")
    print(f"PHASE 1 — {total:,} non-success hospitals  |  Target: {TARGET_NEW_REVIEWS:,}")
    print(f"{'='*70}")

    for i, hosp in enumerate(candidates[start:], start=start):
        if total_new >= TARGET_NEW_REVIEWS:
            print(f"\n[TARGET REACHED] {total_new:,} reviews — stopping Phase 1.")
            break

        name = hosp["name"]
        print(f"\n[PROGRESS] Phase 1: {i+1}/{total} — {name}")
        log.info(f"Phase 1 [{i+1}/{total}]: {name}")

        try:
            added = process_hospital(page, hosp, writer, existing_keys,
                                      PHASE1_MAX_PER_HOSPITAL)
        except Exception as e:
            log.error(f"  Unhandled error for '{name}': {e}")
            log.debug(traceback.format_exc())
            added = 0

        total_new += added
        save_checkpoint("phase1", i + 1, name, total_new)
        print(f"  Total new: {total_new:,} / {TARGET_NEW_REVIEWS:,}")

    print(f"\nPhase 1 done. New reviews this run: {total_new:,}")
    return total_new


# ─────────────────────────────────────────────────────────────────────────────
# Phase 2 — Gap fill on already-successful hospitals
# ─────────────────────────────────────────────────────────────────────────────

def run_phase2(page, hospitals, writer, existing_keys, checkpoint, so_far):
    candidates = [h for h in hospitals if h["collection_status"] not in NON_SUCCESS_STATUSES]
    total, added, start = len(candidates), 0, 0

    if checkpoint and checkpoint.get("phase") == "phase2":
        start = checkpoint.get("hospital_index", 0)
        added = max(0, checkpoint.get("total_new", 0) - so_far)
        print(f"\n[RESUME] Phase 2 from #{start}")

    still_need = TARGET_NEW_REVIEWS - so_far
    print(f"\n{'='*70}")
    print(f"PHASE 2 — {total:,} success hospitals | Still need: {still_need:,}")
    print(f"{'='*70}")

    for i, hosp in enumerate(candidates[start:], start=start):
        if so_far + added >= TARGET_NEW_REVIEWS:
            print(f"\n[TARGET REACHED] {so_far + added:,} total.")
            break

        name = hosp["name"]
        print(f"\n[PROGRESS] Phase 2: {i+1}/{total} — {name}")
        log.info(f"Phase 2 [{i+1}/{total}]: {name}")

        try:
            new = process_hospital(page, hosp, writer, existing_keys,
                                    PHASE2_PER_HOSPITAL)
        except Exception as e:
            log.error(f"  Unhandled error for '{name}': {e}")
            new = 0

        added += new
        save_checkpoint("phase2", i + 1, name, so_far + added)
        print(f"  Total: {so_far + added:,} / {TARGET_NEW_REVIEWS:,}")

    print(f"\nPhase 2 done. Additional reviews: {added:,}")
    return added


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    print("=" * 70)
    print("  HOSPITAL REVIEW AUTOMATION — scraper_final.py v13.0")
    print(f"  Excel : {EXCEL_FILE}")
    print(f"  Target: {TARGET_NEW_REVIEWS:,} new unique reviews")
    print(f"  Browser: {'HEADLESS' if HEADLESS else 'VISIBLE'}")
    print(f"  Start : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    checkpoint = load_checkpoint()
    if checkpoint:
        print(f"\n[CHECKPOINT] phase='{checkpoint.get('phase')}' | "
              f"index={checkpoint.get('hospital_index')} | "
              f"collected={checkpoint.get('total_new', 0):,}")
    else:
        print("\n[FRESH START] No checkpoint found.")

    hospitals = load_hospitals()
    writer = ExcelWriter()
    existing_keys = load_existing_review_keys()
    total_new = 0

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=HEADLESS,
            args=["--disable-dev-shm-usage", "--no-sandbox"],
        )
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/138.0.0.0 Safari/537.36"
            ),
            locale="en-US",
            viewport={"width": 1280, "height": 800},
        )
        page = context.new_page()

        try:
            # Run phases
            skip_phase1 = checkpoint and checkpoint.get("phase") == "phase2"
            if skip_phase1:
                total_new = checkpoint.get("total_new", 0)
                print(f"\n[SKIP] Phase 1 already done ({total_new:,} reviews) — Phase 2…")
            else:
                total_new = run_phase1(page, hospitals, writer, existing_keys, checkpoint)

            if total_new < TARGET_NEW_REVIEWS:
                total_new += run_phase2(page, hospitals, writer, existing_keys,
                                         None if skip_phase1 else checkpoint,
                                         total_new)
            else:
                print(f"\n[TARGET MET] {total_new:,} reviews — Phase 2 not needed.")

        finally:
            browser.close()

    clear_checkpoint()
    print("\n" + "=" * 70)
    print("  SCRAPING COMPLETE")
    print(f"  Total NEW unique reviews : {total_new:,}")
    print(f"  Saved to                 : {EXCEL_FILE}")
    print(f"  Log                      : {_log_file}")
    print(f"  Finished                 : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)


if __name__ == "__main__":
    main()
