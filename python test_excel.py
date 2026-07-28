from openpyxl import load_workbook

wb = load_workbook("HospitalAutomation.xlsx")

print("Workbook loaded successfully.")
print(wb.sheetnames)