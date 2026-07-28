"""
scraper_main.py  v5.0
---------------------
Uses Google search (requests) to find real URLs on all 3 sites.
Fixes: JustDial HTTP2 block, HexaHealth missing API, Practo wrong URLs.
"""

import os, sys, re, json, time, random, hashlib, logging, traceback
from datetime import datetime
from pathlib import Path
from urllib.parse import quote_plus

try:
    import requests
    from bs4 import BeautifulSoup
    from openpyxl import load_workbook
    from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
except ImportError as _e:
    sys.exit(
        f"\n[FATAL] Missing dependency: {_e}\n"
        "Run:  pip install playwright openpyxl beautifulsoup4 requests lxml\n"
        "Then: python -m playwright install chromium\n"
    )

_ROOT      = Path(__file__).resolve().parent
EXCEL_FILE = _ROOT / "output" / "HospitalAutomation.xlsx"
LOG_DIR    = _ROOT / "logs"
CACHE_DIR  = _ROOT / "cache"

LOG_DIR.mkdir(exist_ok=True)
CACHE_DIR.mkdir(exist_ok=True)

CHECKPOINT_FILE = CACHE_DIR / "scraper_main_checkpoint.json"

TARGET_NEW_REVIEWS    = 1500
PHASE1_MAX_PER_SOURCE = 10
PHASE2_PER_HOSPITAL   = 4
SCROLL_ROUNDS         = 2
DELAY_MIN             = 0.8
DELAY_MAX             = 2.0
MAX_RETRIES           = 2
RETRY_BASE_DELAY      = 2
HEADLESS              = True
BROWSER_TIMEOUT       = 20_000
PAGE_WAIT_MS          = 1_500
SCROLL_WAIT_MS        = 800

NON_SUCCESS_STATUSES = {
    "No Match", "No Reviews", "Failed", "Scraper Error",
    "No Review", "Not Found", "Error", "Pending", "Low Reviews", None, "",
}

_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Upgrade-Insecure-Requests": "1",
}

_log_file = LOG_DIR / f"scraper_main_{datetime.now().strftime('%Y%m%d_%H%M%S')}.log"
logging.basicConfig(
    level=logging.INFO,
    format="[%(asctime)s] %(levelname)s | %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(_log_file, encoding="utf-8"),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger("scraper")

def save_checkpoint(phase, index, name, total_new):
    with open(CHECKPOINT_FILE, "w", encoding="utf-8") as f:
        json.dump({"phase": phase, "index": index, "name": name,
                   "total_new": total_new, "saved_at": datetime.now().isoformat()}, f, indent=2)

def load_checkpoint():
    if CHECKPOINT_FILE.exists():
        with open(CHECKPOINT_FILE, encoding="utf-8") as f:
            return json.load(f)
    return None

def clear_checkpoint():
    if CHECKPOINT_FILE.exists():
        CHECKPOINT_FILE.unlink()

def _hash(hospital, reviewer, review_text):
    raw = f"{hospital.strip().lower()}|{reviewer.strip().lower()}|{review_text.strip().lower()}"
    return hashlib.sha256(raw.encode()).hexdigest()

def load_existing_hashes():
    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    ws = wb["Reviews"]
    hashes = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        try:
            h, r, t = str(row[0] or ""), str(row[1] or ""), str(row[3] or "")
            if h or r or t:
                hashes.add(_hash(h, r, t))
        except Exception:
            pass
    wb.close()
    log.info(f"Loaded {len(hashes):,} existing review hashes.")
    return hashes

def load_hospitals():
    if not EXCEL_FILE.exists():
        sys.exit(f"\n[FATAL] Excel not found: {EXCEL_FILE}")
    wb = load_workbook(EXCEL_FILE, read_only=True, data_only=True)
    if "Hospitals" not in wb.sheetnames:
        sys.exit(f"\n[FATAL] Sheet 'Hospitals' missing. Found: {wb.sheetnames}")
    ws  = wb["Hospitals"]
    out = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[1] if len(row) > 1 else None
        if not name:
            continue
        out.append({
            "name":              str(name).strip(),
            "collection_status": str(row[8] if len(row) > 8 else "").strip(),
            "reviews_found":     (row[7] if len(row) > 7 else 0) or 0,
        })
    wb.close()
    log.info(f"Loaded {len(out):,} hospitals.")
    return out

def append_reviews(new_rows):
    if not new_rows:
        return
    wb = load_workbook(EXCEL_FILE)
    ws = wb["Reviews"]
    for r in new_rows:
        ws.append([
            r.get("hospital_name", ""),
            r.get("reviewer",      "Unknown"),
            r.get("rating",        0),
            r.get("review",        ""),
            r.get("date",          ""),
            r.get("source",        ""),
            "",
        ])
    wb.save(EXCEL_FILE)
    wb.close()

def update_hospital_status(hospital_name, reviews_added, status):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Hospitals"]
        for row in ws.iter_rows(min_row=2):
            if str(row[1].value or "").strip() == hospital_name:
                row[7].value = reviews_added
                row[8].value = status
                break
        wb.save(EXCEL_FILE)
        wb.close()
    except Exception as e:
        log.warning(f"Could not update status for '{hospital_name}': {e}")

_sess = requests.Session()
_sess.headers.update(_HEADERS)

def _http_get(url, timeout=12):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            r = _sess.get(url, timeout=timeout, allow_redirects=True)
            if r.status_code == 200:
                return r.text
            return None
        except requests.exceptions.Timeout:
            log.debug(f"  HTTP timeout attempt {attempt}: {url}")
        except Exception as e:
            log.debug(f"  HTTP error attempt {attempt}: {e}")
        if attempt < MAX_RETRIES:
            time.sleep(RETRY_BASE_DELAY)
    return None

def _google_find_url(hospital_name, site):
    query = f"{hospital_name} reviews site:{site}"
    goog  = f"https://www.google.com/search?q={quote_plus(query)}&num=5&hl=en"
    try:
        r = requests.get(goog, headers=_HEADERS, timeout=10, allow_redirects=True)
        if r.status_code != 200:
            log.debug(f"  Google {r.status_code} for {site}")
            return None
        soup = BeautifulSoup(r.text, "lxml")
        for a in soup.select("a[href]"):
            href = a.get("href", "")
            m = re.search(r"/url\?q=(https?://[^&]+)", href)
            if m:
                href = m.group(1)
            if (site in href and
                    not any(x in href for x in ["google.com", "googleadservices", "webcache"])):
                href = href.split("&")[0] if "&" in href else href
                return href
    except Exception as e:
        log.debug(f"  Google search error ({site}): {e}")
    return None

def _delay():
    time.sleep(random.uniform(DELAY_MIN, DELAY_MAX))

def _safe_goto(page, url):
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            page.goto(url, wait_until="domcontentloaded", timeout=BROWSER_TIMEOUT)
            return True
        except PWTimeout:
            log.warning(f"  Browser timeout: {url} (attempt {attempt}/{MAX_RETRIES})")
        except Exception as e:
            log.warning(f"  Browser nav error: {e} (attempt {attempt}/{MAX_RETRIES})")
        if attempt < MAX_RETRIES:
            time.sleep(RETRY_BASE_DELAY)
    return False

def _scroll_and_load(page):
    load_more_sels = [
        "button:has-text('Load More')",  "button:has-text('Show More')",
        "button:has-text('View More')",  "a:has-text('Load More')",
        "[class*='loadMore']",           "[class*='load-more']",
    ]
    for _ in range(SCROLL_ROUNDS):
        try:
            page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            page.wait_for_timeout(SCROLL_WAIT_MS)
            for sel in load_more_sels:
                try:
                    btn = page.locator(sel).first
                    if btn.count() and btn.is_visible():
                        btn.scroll_into_view_if_needed()
                        btn.click()
                        page.wait_for_timeout(SCROLL_WAIT_MS)
                        break
                except Exception:
                    pass
        except Exception:
            pass

def _browser_get_html(page, url):
    if not _safe_goto(page, url):
        return None
    page.wait_for_timeout(PAGE_WAIT_MS)
    _scroll_and_load(page)
    try:
        return page.content()
    except Exception:
        return None

def _parse_hexahealth(html, max_n):
    soup  = BeautifulSoup(html, "lxml")
    cards = (
        soup.select("div[class*='ReviewCard']") or
        soup.select("div[class*='review-card']") or
        soup.select("div[class*='reviewCard']") or
        soup.select("div[class*='review_card']") or
        soup.select("div[class*='patientReview']") or
        soup.select("div[class*='patient-review']") or
        [d for d in soup.select("div")
         if len((d.select_one("p") or {}).get_text(strip=True)
                if d.select_one("p") else "") > 30
         and (d.select("img[src*='star']") or d.select("[class*='star']"))]
    )
    out  = []
    seen = set()
    for card in cards:
        if len(out) >= max_n:
            break
        p = (card.select_one("[class*='review-text']") or
             card.select_one("[class*='reviewText']") or
             card.select_one("[class*='review_text']") or
             card.select_one("[class*='comment']") or
             card.select_one("p"))
        text = p.get_text(strip=True) if p else ""
        if not text or len(text) < 20 or text in seen:
            continue
        seen.add(text)
        ntag = (card.select_one("[class*='reviewer']") or
                card.select_one("[class*='author']") or
                card.select_one("span[class*='name']") or
                card.select_one("strong"))
        reviewer = ntag.get_text(strip=True) if ntag else "Unknown"
        rating   = len(card.select(
            "img[src*='star_yellow'], img[src*='staryellow'], "
            "[class*='filled'], [class*='active'][class*='star']"
        ))
        date = ""
        for s in reversed(card.select("span")):
            st = s.get_text(strip=True)
            if st and re.search(r"\d{4}|\d+ (day|week|month|year|ago)", st, re.I):
                date = st
                break
        out.append({"reviewer": reviewer, "rating": rating, "review": text, "date": date})
    return out

def _parse_practo(html, max_n):
    soup  = BeautifulSoup(html, "lxml")
    cards = (
        soup.select("div[data-qa-id='review_card']") or
        soup.select("div[class*='doctor-review']") or
        soup.select("div[class*='review-card']") or
        soup.select("div[class*='ReviewCard']") or
        soup.select("div[class*='reviews-item']") or
        soup.select("li[class*='review']") or
        [d for d in soup.select("div")
         if len((d.select_one("p") or {}).get_text(strip=True)
                if d.select_one("p") else "") > 30
         and d.select("[class*='star'], [class*='rating']")]
    )
    out  = []
    seen = set()
    for card in cards:
        if len(out) >= max_n:
            break
        ntag = (card.find(attrs={"data-qa-id": "reviewer_name"}) or
                card.find(class_=re.compile(r"reviewer|author|user.?name", re.I)) or
                card.find("strong"))
        reviewer = ntag.get_text(strip=True) if ntag else "Unknown"
        filled   = card.select(
            "i.c-icon--star-filled, span[class*='star-fill'], "
            "i[class*='star'][class*='fill'], img[alt*='star'][src*='fill']"
        )
        rating = len(filled)
        if rating == 0:
            rtag = (card.find(attrs={"data-qa-id": "star_rating"}) or
                    card.find(class_=re.compile(r"rating|score", re.I)))
            if rtag:
                m = re.search(r"\d+\.?\d*", rtag.get_text())
                rating = float(m.group()) if m else 0
        ttag = (card.find(attrs={"data-qa-id": "review_text"}) or
                card.find(class_=re.compile(r"review.?text|comment|body|content", re.I)) or
                card.find("p"))
        text = ttag.get_text(strip=True) if ttag else ""
        if not text or len(text) < 20 or text in seen:
            continue
        seen.add(text)
        dtag = card.find(class_=re.compile(r"date|time|when|ago", re.I))
        date = dtag.get_text(strip=True) if dtag else ""
        out.append({"reviewer": reviewer, "rating": rating, "review": text, "date": date})
    return out

def _parse_justdial(html, max_n):
    soup  = BeautifulSoup(html, "lxml")
    cards = (
        soup.select("div[class*='jd-review']") or
        soup.select("div[class*='reviewWrap']") or
        soup.select("div[class*='review-item']") or
        soup.select("li[class*='review']") or
        soup.select("div[class*='review']") or
        [d for d in soup.select("div")
         if len((d.select_one("p") or {}).get_text(strip=True)
                if d.select_one("p") else "") > 30
         and d.select("[class*='star'], [class*='rating'], [class*='fill']")]
    )
    out  = []
    seen = set()
    for card in cards:
        if len(out) >= max_n:
            break
        ntag = (card.find(class_=re.compile(r"reviewer|username|profile|user-name", re.I)) or
                card.find("strong") or card.find("b"))
        reviewer = ntag.get_text(strip=True) if ntag else "Unknown"
        filled   = card.select(
            "span[class*='fill'], i[class*='fill'], "
            "span[class*='active'], img[src*='ratingstar_full']"
        )
        rating = len(filled) if filled else 0
        ttag   = (card.find(class_=re.compile(r"review.?text|review.?detail|comment|desc|content", re.I)) or
                  card.find("p"))
        text   = ttag.get_text(strip=True) if ttag else ""
        if not text or len(text) < 20 or text in seen:
            continue
        seen.add(text)
        dtag = card.find(class_=re.compile(r"date|time|ago|when|posted", re.I))
        date = dtag.get_text(strip=True) if dtag else ""
        out.append({"reviewer": reviewer, "rating": rating, "review": text, "date": date})
    return out

_PARSERS = {
    "hexahealth.com": _parse_hexahealth,
    "practo.com":     _parse_practo,
    "justdial.com":   _parse_justdial,
}

def scrape_site(page, site, hospital_name, max_n):
    parse = _PARSERS[site]

    url = _google_find_url(hospital_name, site)
    if not url:
        log.debug(f"  {site}: Google found no URL for '{hospital_name}'")
        return []

    url_clean = url.split("?")[0].rstrip("/")
    if site in ("hexahealth.com", "practo.com") and not url_clean.endswith("/reviews"):
        url_clean += "/reviews"

    log.info(f"  {site} URL: {url_clean}")

    html = _http_get(url_clean)
    if html:
        rows = parse(html, max_n)
        if rows:
            log.info(f"  {site} (HTTP): {len(rows)} reviews parsed")
            return rows
        log.debug(f"  {site}: HTTP OK but 0 reviews — trying browser")
    else:
        log.debug(f"  {site}: HTTP failed — trying browser")

    browser_html = _browser_get_html(page, url_clean)
    if not browser_html:
        return []
    rows = parse(browser_html, max_n)
    log.info(f"  {site} (browser): {len(rows)} reviews parsed")
    return rows

SOURCES = ["hexahealth.com", "practo.com", "justdial.com"]

def process_hospital(page, hospital, existing_hashes, max_per_source):
    name     = hospital["name"]
    new_rows = []
    for site in SOURCES:
        try:
            raw       = scrape_site(page, site, name, max_per_source)
            added_now = 0
            for r in raw:
                h = _hash(name, r.get("reviewer", ""), r.get("review", ""))
                if h in existing_hashes or not r.get("review", "").strip():
                    continue
                existing_hashes.add(h)
                new_rows.append({
                    "hospital_name": name,
                    "reviewer":      r.get("reviewer", "Unknown"),
                    "rating":        r.get("rating",   0),
                    "review":        r.get("review",   ""),
                    "date":          r.get("date",     ""),
                    "source":        site,
                })
                added_now += 1
            if raw:
                log.info(f"  {site}: {len(raw)} scraped → {added_now} new unique")
            _delay()
        except Exception as e:
            log.error(f"  Error on {site} for '{name}': {e}")
            log.debug(traceback.format_exc())
            _delay()
    return new_rows

def run_phase1(page, hospitals, hashes, checkpoint):
    candidates = [h for h in hospitals if h["collection_status"] in NON_SUCCESS_STATUSES]
    total      = len(candidates)
    total_new  = 0
    start      = 0

    if checkpoint and checkpoint.get("phase") == "phase1":
        start     = checkpoint.get("index", 0)
        total_new = checkpoint.get("total_new", 0)
        print(f"\n[RESUME] Phase 1 from #{start} ({checkpoint.get('name','')}) | Collected: {total_new:,}")

    print(f"\n{'='*70}")
    print(f"PHASE 1 — {total:,} non-success hospitals")
    print(f"Target: {TARGET_NEW_REVIEWS:,} new unique reviews")
    print(f"{'='*70}")

    for i, hosp in enumerate(candidates[start:], start=start):
        if total_new >= TARGET_NEW_REVIEWS:
            print(f"\n[TARGET REACHED] {total_new:,} — stopping Phase 1 early.")
            break
        name = hosp["name"]
        print(f"\n[Phase 1] {i+1}/{total} — {name}")
        log.info(f"Phase 1 [{i+1}/{total}]: {name}")
        rows = process_hospital(page, hosp, hashes, PHASE1_MAX_PER_SOURCE)
        if rows:
            append_reviews(rows)
            total_new += len(rows)
            update_hospital_status(name, len(rows), "Success")
            print(f"  ✓ +{len(rows)} | Total: {total_new:,} / {TARGET_NEW_REVIEWS:,}")
        else:
            update_hospital_status(name, 0, "No Reviews")
            print(f"  — No reviews found.")
        save_checkpoint("phase1", i + 1, name, total_new)

    print(f"\nPhase 1 done. New reviews: {total_new:,}")
    return total_new

def run_phase2(page, hospitals, hashes, checkpoint, so_far):
    candidates = [h for h in hospitals if h["collection_status"] not in NON_SUCCESS_STATUSES]
    total  = len(candidates)
    added  = 0
    start  = 0

    if checkpoint and checkpoint.get("phase") == "phase2":
        start = checkpoint.get("index", 0)
        added = max(0, checkpoint.get("total_new", 0) - so_far)
        print(f"\n[RESUME] Phase 2 from #{start} ({checkpoint.get('name','')})")

    print(f"\n{'='*70}")
    print(f"PHASE 2 — {total:,} success hospitals (up to {PHASE2_PER_HOSPITAL} more each)")
    print(f"Still need: {TARGET_NEW_REVIEWS - so_far:,} reviews")
    print(f"{'='*70}")

    for i, hosp in enumerate(candidates[start:], start=start):
        if so_far + added >= TARGET_NEW_REVIEWS:
            print(f"\n[TARGET REACHED] {so_far + added:,} total — stopping Phase 2.")
            break
        name = hosp["name"]
        print(f"\n[Phase 2] {i+1}/{total} — {name}")
        log.info(f"Phase 2 [{i+1}/{total}]: {name}")
        rows = process_hospital(page, hosp, hashes, PHASE2_PER_HOSPITAL)
        if rows:
            append_reviews(rows)
            added += len(rows)
            print(f"  ✓ +{len(rows)} | Total: {so_far + added:,} / {TARGET_NEW_REVIEWS:,}")
        else:
            print(f"  — No new reviews.")
        save_checkpoint("phase2", i + 1, name, so_far + added)

    print(f"\nPhase 2 done. Additional reviews: {added:,}")
    return added

def main():
    print("=" * 70)
    print("  HOSPITAL REVIEW AUTOMATION — scraper_main.py v5.0")
    print(f"  Excel : {EXCEL_FILE}")
    print(f"  Target: {TARGET_NEW_REVIEWS:,} new unique reviews")
    print(f"  Start : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)

    checkpoint = load_checkpoint()
    if checkpoint:
        print(f"\n[CHECKPOINT] phase='{checkpoint['phase']}' | "
              f"index={checkpoint['index']} | collected={checkpoint['total_new']:,}")
    else:
        print("\n[FRESH START] No checkpoint found.")

    hospitals = load_hospitals()
    hashes    = load_existing_hashes()
    total_new = 0

    with sync_playwright() as pw:
        browser = pw.chromium.launch(
            headless=HEADLESS,
            args=[
                "--disable-gpu",
                "--disable-dev-shm-usage",
                "--no-sandbox",
                "--disable-extensions",
                "--disable-background-networking",
                "--memory-pressure-off",
            ],
        )
        context = browser.new_context(
            user_agent=_HEADERS["User-Agent"],
            locale="en-US",
            viewport={"width": 1280, "height": 800},
        )
        context.route(
            "**/*.{png,jpg,jpeg,gif,webp,svg,ico,woff,woff2,ttf,eot,mp4,mp3}",
            lambda route: route.abort()
        )
        page = context.new_page()

        if not checkpoint or checkpoint.get("phase") == "phase1":
            total_new = run_phase1(page, hospitals, hashes, checkpoint)
        elif checkpoint.get("phase") == "phase2":
            total_new = checkpoint.get("total_new", 0)
            print(f"\n[SKIP] Phase 1 done ({total_new:,} reviews) — going to Phase 2…")

        if total_new < TARGET_NEW_REVIEWS:
            total_new += run_phase2(page, hospitals, hashes, checkpoint, total_new)
        else:
            print(f"\n[TARGET MET] {total_new:,} reviews — Phase 2 not needed.")

        browser.close()

    clear_checkpoint()

    print("\n" + "=" * 70)
    print("  SCRAPING COMPLETE")
    print(f"  Total NEW unique reviews : {total_new:,}")
    print(f"  Saved to                 : {EXCEL_FILE}")
    print(f"  Finished                 : {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 70)
    log.info(f"Done — {total_new} new reviews added.")

if __name__ == "__main__":
    main()